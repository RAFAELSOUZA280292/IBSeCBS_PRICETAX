"""
test_batch_excel.py
===================
Teste local da geração do Excel com duplicação de linhas para NCMs ambíguos.
Usa os XMLs reais do ZIP enviado pelo usuário.

Execução:
    python3 test_batch_excel.py
"""

import sys
import os
import glob

# Garantir que o diretório do projeto está no path
sys.path.insert(0, os.path.dirname(__file__))

from batch_xml_processor import (
    process_single_xml,
    generate_excel_report,
    _build_dados_completos_expandidos,
)

# -----------------------------------------------------------------------
# 1. Processar os XMLs reais
# -----------------------------------------------------------------------
XML_DIR = "/tmp/xml_analise"
xml_files = sorted(glob.glob(os.path.join(XML_DIR, "*.xml")))

print(f"[TESTE] Encontrados {len(xml_files)} XMLs em {XML_DIR}")
assert len(xml_files) > 0, "Nenhum XML encontrado!"

resultados = []
for xml_path in xml_files:
    r = process_single_xml(xml_path, save_to_collector=False)
    resultados.append(r)
    print(f"  [{r['status']:10s}] {r['arquivo'][:50]} — {r['total_itens']} itens")

total_itens = sum(r['total_itens'] for r in resultados)
print(f"\n[TESTE] Total de itens processados: {total_itens}")

# -----------------------------------------------------------------------
# 2. Testar _build_dados_completos_expandidos
# -----------------------------------------------------------------------
linhas = _build_dados_completos_expandidos(resultados)
print(f"[TESTE] Linhas expandidas geradas: {len(linhas)}")

# Verificar que NCMs ambíguos foram duplicados
ncms_ambiguos = {}
for linha in linhas:
    tipo = linha.get('Tipo Linha', '')
    ncm  = linha.get('NCM', '')
    if 'OPCAO' in tipo:
        ncms_ambiguos[ncm] = ncms_ambiguos.get(ncm, 0) + 1

print(f"[TESTE] NCMs ambíguos expandidos: {dict(ncms_ambiguos)}")

# Verificar que NCMs únicos têm Tipo Linha = UNICO
linhas_unicas = [l for l in linhas if l.get('Tipo Linha') == 'UNICO']
print(f"[TESTE] Linhas com tratamento único: {len(linhas_unicas)}")

# Verificar que coluna ALERTA existe em todas as linhas
sem_alerta = [l for l in linhas if 'ALERTA' not in l]
assert len(sem_alerta) == 0, f"ERRO: {len(sem_alerta)} linhas sem coluna ALERTA!"
print("[TESTE] Coluna ALERTA presente em todas as linhas: OK")

# -----------------------------------------------------------------------
# 3. Gerar o Excel completo
# -----------------------------------------------------------------------
print("\n[TESTE] Gerando Excel...")
excel_bytes = generate_excel_report(resultados)
assert excel_bytes is not None
assert excel_bytes.getbuffer().nbytes > 0

# Salvar para inspeção manual
output_path = "/tmp/test_relatorio_lote.xlsx"
with open(output_path, "wb") as f:
    f.write(excel_bytes.read())

print(f"[TESTE] Excel gerado com sucesso: {output_path}")
print(f"[TESTE] Tamanho: {os.path.getsize(output_path):,} bytes")

# -----------------------------------------------------------------------
# 4. Validar estrutura do Excel
# -----------------------------------------------------------------------
import openpyxl
wb = openpyxl.load_workbook(output_path)
print(f"\n[TESTE] Abas do Excel: {wb.sheetnames}")

assert 'Resumo' in wb.sheetnames,          "ERRO: Aba 'Resumo' não encontrada!"
assert 'Validacao' in wb.sheetnames,       "ERRO: Aba 'Validacao' não encontrada!"
assert 'Dados Completos' in wb.sheetnames, "ERRO: Aba 'Dados Completos' não encontrada!"

ws_dc = wb['Dados Completos']
headers = [cell.value for cell in ws_dc[1]]
print(f"[TESTE] Colunas 'Dados Completos': {headers[:6]}...")

assert 'ALERTA' in headers,       "ERRO: Coluna ALERTA não encontrada!"
assert 'Tipo Linha' in headers,   "ERRO: Coluna 'Tipo Linha' não encontrada!"
assert 'cClassTrib' in headers,   "ERRO: Coluna 'cClassTrib' não encontrada!"
assert 'Situacao Operacao' in headers, "ERRO: Coluna 'Situacao Operacao' não encontrada!"

total_linhas_dc = ws_dc.max_row - 1  # excluir cabeçalho
print(f"[TESTE] Total de linhas em 'Dados Completos': {total_linhas_dc}")
assert total_linhas_dc == len(linhas), \
    f"ERRO: Esperado {len(linhas)} linhas, encontrado {total_linhas_dc}!"

# Verificar formatação: célula ALERTA da primeira linha de dados deve ter fill
idx_alerta = headers.index('ALERTA') + 1
cell_alerta = ws_dc.cell(row=2, column=idx_alerta)
assert cell_alerta.fill.fgColor.rgb not in ('00000000', 'FFFFFFFF', None), \
    "ERRO: Coluna ALERTA sem formatação de cor!"
print(f"[TESTE] Formatação coluna ALERTA: cor={cell_alerta.fill.fgColor.rgb} — OK")

# Verificar cabeçalho com cor PRICETAX
header_cell = ws_dc.cell(row=1, column=1)
assert header_cell.fill.fgColor.rgb == 'FF1A1A1A', \
    f"ERRO: Cabeçalho sem cor preta PRICETAX! Cor atual: {header_cell.fill.fgColor.rgb}"
print(f"[TESTE] Formatação cabeçalho: cor={header_cell.fill.fgColor.rgb} — OK")

print("\n[TESTE] ===== TODOS OS TESTES PASSARAM ===== ")
print(f"[TESTE] Excel disponível para inspeção em: {output_path}")

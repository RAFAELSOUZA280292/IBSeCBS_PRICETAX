"""
PRICETAX - Exportador Excel Profissional
=========================================

Módulo para exportar DataFrames para Excel (.xlsx) com formatação premium
no padrão visual PRICETAX (amarelo + preto).

Autor: PRICETAX
Data: 08/02/2026
"""

from io import BytesIO
from typing import Optional, List, Dict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet


# Cores PRICETAX
COLOR_GOLD = "FFDD00"  # Amarelo PRICETAX
COLOR_BLACK = "000000"
COLOR_GRAY_DARK = "2A2A2A"
COLOR_GRAY_LIGHT = "F5F5F5"
COLOR_WHITE = "FFFFFF"


def aplicar_estilo_header(ws: Worksheet, max_col: int):
    """
    Aplica estilo profissional ao cabeçalho (primeira linha).
    
    Args:
        ws: Worksheet do openpyxl
        max_col: Número máximo de colunas
    """
    # Estilo do header
    header_font = Font(name='Poppins', size=11, bold=True, color=COLOR_BLACK)
    header_fill = PatternFill(start_color=COLOR_GOLD, end_color=COLOR_GOLD, fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin', color=COLOR_BLACK),
        right=Side(style='thin', color=COLOR_BLACK),
        top=Side(style='thin', color=COLOR_BLACK),
        bottom=Side(style='thin', color=COLOR_BLACK)
    )
    
    # Aplicar estilo em todas as células do header
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border


def aplicar_estilo_dados(ws: Worksheet, max_row: int, max_col: int):
    """
    Aplica estilo profissional aos dados (linhas alternadas).
    
    Args:
        ws: Worksheet do openpyxl
        max_row: Número máximo de linhas
        max_col: Número máximo de colunas
    """
    # Estilos
    data_font = Font(name='Poppins', size=10, color=COLOR_BLACK)
    data_alignment = Alignment(horizontal='left', vertical='center')
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Cores alternadas
    fill_white = PatternFill(start_color=COLOR_WHITE, end_color=COLOR_WHITE, fill_type="solid")
    fill_gray = PatternFill(start_color=COLOR_GRAY_LIGHT, end_color=COLOR_GRAY_LIGHT, fill_type="solid")
    
    # Aplicar estilo em todas as células de dados
    for row in range(2, max_row + 1):
        # Alternar cores
        fill = fill_gray if row % 2 == 0 else fill_white
        
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border
            cell.fill = fill


def ajustar_largura_colunas(ws: Worksheet):
    """
    Ajusta automaticamente a largura das colunas baseado no conteúdo.
    
    Args:
        ws: Worksheet do openpyxl
    """
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        # Largura ajustada (min 12, max 50)
        adjusted_width = min(max(max_length + 2, 12), 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def formatar_valores_monetarios(ws: Worksheet, colunas_monetarias: List[str]):
    """
    Formata colunas monetárias com padrão brasileiro (R$ 1.234,56).
    
    Args:
        ws: Worksheet do openpyxl
        colunas_monetarias: Lista de nomes de colunas que contêm valores monetários
    """
    # Encontrar índices das colunas monetárias
    header_row = [cell.value for cell in ws[1]]
    indices_monetarios = []
    
    for col_name in colunas_monetarias:
        if col_name in header_row:
            indices_monetarios.append(header_row.index(col_name) + 1)
    
    # Aplicar formato monetário
    for col_idx in indices_monetarios:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value and isinstance(cell.value, (int, float)):
                cell.number_format = 'R$ #,##0.00'


def exportar_dataframe_para_excel(
    df: pd.DataFrame,
    nome_aba: str = "Dados",
    colunas_monetarias: Optional[List[str]] = None
) -> BytesIO:
    """
    Exporta DataFrame para Excel com formatação profissional PRICETAX.
    
    Args:
        df: DataFrame a ser exportado
        nome_aba: Nome da aba no Excel
        colunas_monetarias: Lista de colunas que contêm valores monetários
        
    Returns:
        BytesIO com o arquivo Excel
    """
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = nome_aba[:31]  # Limite de 31 caracteres para nome de aba
    
    # Escrever dados
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # Aplicar estilos
    aplicar_estilo_header(ws, ws.max_column)
    aplicar_estilo_dados(ws, ws.max_row, ws.max_column)
    ajustar_largura_colunas(ws)
    
    # Formatar valores monetários
    if colunas_monetarias:
        formatar_valores_monetarios(ws, colunas_monetarias)
    
    # Adicionar filtro automático
    ws.auto_filter.ref = ws.dimensions
    
    # Congelar primeira linha (header)
    ws.freeze_panes = 'A2'
    
    # Salvar em BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer


def exportar_multiplas_abas(
    dataframes: Dict[str, pd.DataFrame],
    colunas_monetarias_por_aba: Optional[Dict[str, List[str]]] = None
) -> BytesIO:
    """
    Exporta múltiplos DataFrames para Excel com múltiplas abas formatadas.
    
    Args:
        dataframes: Dicionário {nome_aba: DataFrame}
        colunas_monetarias_por_aba: Dicionário {nome_aba: [colunas_monetarias]}
        
    Returns:
        BytesIO com o arquivo Excel
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remover aba padrão
    
    for nome_aba, df in dataframes.items():
        # Criar nova aba
        ws = wb.create_sheet(title=nome_aba[:31])
        
        # Escrever dados
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Aplicar estilos
        aplicar_estilo_header(ws, ws.max_column)
        aplicar_estilo_dados(ws, ws.max_row, ws.max_column)
        ajustar_largura_colunas(ws)
        
        # Formatar valores monetários
        if colunas_monetarias_por_aba and nome_aba in colunas_monetarias_por_aba:
            formatar_valores_monetarios(ws, colunas_monetarias_por_aba[nome_aba])
        
        # Adicionar filtro automático
        ws.auto_filter.ref = ws.dimensions
        
        # Congelar primeira linha
        ws.freeze_panes = 'A2'
    
    # Salvar em BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer

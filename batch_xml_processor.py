"""
batch_xml_processor.py
======================
Processamento em Lote de XMLs NFe — PRICETAX

Responsabilidades deste módulo:
    1. Validar e processar até 500 XMLs de NFe simultaneamente.
    2. Enriquecer cada item com cClassTrib, alíquotas IBS/CBS e base legal
       via motor tributário (calcular_tributacao_completa + BeneficiosFiscaisEngine).
    3. Identificar NCMs ambíguos (mais de um cClassTrib possível conforme
       cclasstrib_mapping.py) e sinalizá-los com flag requer_revisao_manual=True.
    4. Gerar relatório Excel profissional com 4 abas:
       - Resumo: estatísticas gerais + NCMs ambíguos identificados
       - Validacao: lista de todos os XMLs com status (CONFORME/DIVERGENTE/ERRO)
       - Divergencias: apenas XMLs com problemas
       - Dados Completos: detalhamento por item com DUPLICACAO DE LINHAS para
         NCMs ambíguos (uma linha por cClassTrib possível) e formatacao visual
         diferenciada (laranja=ambíguo, verde=único).

Lógica de duplicação de linhas (NCMs ambíguos):
    - NCM ambíguo: gera N linhas (uma por cClassTrib possível) com coluna
      'Tipo Linha' = 'OPCAO X DE N' e fundo laranja no Excel.
    - NCM único: gera 1 linha com 'Tipo Linha' = 'UNICO' e fundo verde.
    - O usuário deve excluir as linhas que não se aplicam antes de importar
      no sistema fiscal.

Dependencias principais:
    - calcular_tributacao.py  : calcular_tributacao_completa()
    - tributacao.py           : guess_cclasstrib(), get_class_info_by_code()
    - beneficios_fiscais.py   : BeneficiosFiscaisEngine, consulta_ncm, init_engine
    - cclasstrib_mapping.py   : get_opcoes_cclasstrib_por_ncm(), ncm_tem_multiplos_cclasstrib()
    - xml_parser.py           : parse_nfe_xml()
    - data_collector.py       : save_nfe_data()
    - BDBENEF_PRICETAX_2026.xlsx : planilha de benefícios fiscais

Autor: PRICETAX
Data: 06/02/2026
Atualizado: 20/03/2026 — Duplicação de linhas para NCMs ambíguos + correção de imports
"""

import os
import zipfile
import tempfile
from typing import List, Dict, Any, Tuple, Optional
from datetime import datetime
import pandas as pd
from io import BytesIO
from logger_config import get_logger

# Configurar logging
logger = get_logger(__name__)

# Importar módulos existentes
from xml_parser import parse_nfe_xml
from data_collector import save_nfe_data

# ---------------------------------------------------------------------------
# Importar motor tributário IBS/CBS
# ---------------------------------------------------------------------------
# calcular_tributacao_completa: função centralizada de cálculo (calcular_tributacao.py)
# guess_cclasstrib / get_class_info_by_code: funções de classificação (tributacao.py)
# BeneficiosFiscaisEngine / consulta_ncm: motor de benefícios fiscais (beneficios_fiscais.py)
# init_engine: inicializador do engine com caminho da planilha
# ---------------------------------------------------------------------------
try:
    from calcular_tributacao import calcular_tributacao_completa  # modulo correto
    from tributacao import guess_cclasstrib, get_class_info_by_code
    from beneficios_fiscais import BeneficiosFiscaisEngine, consulta_ncm, init_engine
    import os as _os
    from pathlib import Path as _Path

    # Localizar a planilha de benefícios fiscais (mesmo caminho usado pelo app.py)
    _planilha_paths = [
        _Path("BDBENEF_PRICETAX_2026.xlsx"),
        _Path(_os.getcwd()) / "BDBENEF_PRICETAX_2026.xlsx",
    ]
    try:
        _planilha_paths.append(_Path(__file__).parent / "BDBENEF_PRICETAX_2026.xlsx")
    except Exception:
        pass

    _planilha_encontrada = None
    for _p in _planilha_paths:
        if _p.exists():
            _planilha_encontrada = str(_p)
            break

    if _planilha_encontrada:
        _BENEFICIOS_ENGINE = init_engine(_planilha_encontrada)
        TRIB_DISPONIVEL = True
        logger.info(f"Motor tributário inicializado com planilha: {_planilha_encontrada}")
    else:
        logger.warning("Planilha BDBENEF_PRICETAX_2026.xlsx não encontrada. Motor tributário desabilitado.")
        _BENEFICIOS_ENGINE = None
        TRIB_DISPONIVEL = False
except Exception as _e:
    logger.warning(f"Motor tributário não disponível: {_e}")
    TRIB_DISPONIVEL = False
    _BENEFICIOS_ENGINE = None


def enriquecer_item_tributario(ncm: str, cfop: str) -> dict:
    """
    Enriquece um item de NFe com dados tributários IBS/CBS completos.

    Fluxo interno:
        1. Limpa o NCM (remove não-dígitos).
        2. Consulta cclasstrib_mapping.get_opcoes_cclasstrib_por_ncm() para
           verificar se o NCM possui mais de um cClassTrib possível.
           Se sim, sinaliza requer_revisao_manual=True e monta opcoes_cclasstrib.
        3. Chama calcular_tributacao_completa() com o BeneficiosFiscaisEngine
           para obter regime, alíquotas e cClassTrib sugerido.
        4. Extrai base legal e anexo do primeiro enquadramento de benefício.
        5. Em caso de falha (motor indisponível ou exceção), retorna fallback
           com alíquota padrão 1,025% (IBS 0,125% + CBS 0,9%).

    Args:
        ncm  : Código NCM do produto (8 dígitos, pode conter pontos/hífens).
        cfop : CFOP da operação fiscal (ex: '5405', '6655').

    Returns:
        Dict com as seguintes chaves:
            cclasstrib_code      : str  — código cClassTrib sugerido (ex: '620001')
            cclasstrib_msg       : str  — descrição do cClassTrib
            regime               : str  — regime IVA (ex: 'MONOFASICO', 'RED_60_ESSENCIALIDADE')
            reducao_pct          : int  — percentual de redução de base (0, 30 ou 60)
            ibs_uf               : float — alíquota IBS estadual (%)
            ibs_mun              : float — alíquota IBS municipal (%)
            cbs                  : float — alíquota CBS (%)
            total_iva            : float — alíquota total IVA (%)
            anexo                : str  — anexo da LC 214/2025 (ex: 'ANEXO IV')
            descricao_beneficio  : str  — descrição do benefício fiscal
            base_legal           : str  — artigo da LC 214/2025
            requer_revisao_manual: bool — True se NCM tem múltiplos cClassTribs
            opcoes_cclasstrib    : str  — opções separadas por ' | ' (se ambíguo)
    """
    # Fallback seguro: alíquota padrão LC 214/2025 (ano-teste 2026)
    # IBS UF: 0,10% | IBS Municipal: 0,025% | CBS: 0,90% | Total: 1,025%
    fallback = {
        'cclasstrib_code': '000001',
        'cclasstrib_msg': 'Operação tributada integralmente',
        'regime': 'TRIBUTACAO_PADRAO',
        'reducao_pct': 0,
        'ibs_uf': 0.10,
        'ibs_mun': 0.025,
        'cbs': 0.90,
        'total_iva': 1.025,
        'anexo': '',
        'descricao_beneficio': '',
        'base_legal': 'LC 214/2025 — Alíquota padrão'
    }

    # Se o motor tributário não está disponível (planilha não encontrada),
    # retorna o fallback imediatamente sem tentar calcular.
    if not TRIB_DISPONIVEL:
        fallback['requer_revisao_manual'] = False
        fallback['opcoes_cclasstrib'] = ''
        return fallback

    try:
        import re
        from cclasstrib_mapping import get_opcoes_cclasstrib_por_ncm

        # Passo 1: Normalizar NCM (remover pontos, hífens e espaços)
        ncm_clean = re.sub(r'\D+', '', ncm)

        # Passo 2: Verificar ambiguidade de cClassTrib para este NCM
        # NCMs ambíguos são aqueles com mais de uma classificação tributária possível
        # conforme a LC 214/2025 (ex: gasolina pode ser 620001, 620004, 620005 ou 620006
        # dependendo do percentual de EAC e da posição na cadeia de distribuição).
        opcoes = get_opcoes_cclasstrib_por_ncm(ncm_clean)
        requer_revisao = len(opcoes) > 1
        # Montar string de opções para exibição no Excel (coluna 'opcoes_cclasstrib')
        opcoes_str = ' | '.join([f"{op['code']}: {op['situacao']}" for op in opcoes]) if requer_revisao else ''

        # Passo 3: Calcular tributação completa via motor PRICETAX
        resultado = calcular_tributacao_completa(
            ncm=ncm_clean,
            cfop=cfop,
            beneficios_engine=_BENEFICIOS_ENGINE,
            consulta_ncm_func=consulta_ncm,
            guess_cclasstrib_func=guess_cclasstrib,
            get_class_info_func=get_class_info_by_code,
        )

        # Passo 4: Extrair base legal e anexo do primeiro enquadramento de benefício
        # (quando há múltiplos enquadramentos, usa o de maior prioridade = índice 0)
        base_legal = 'LC 214/2025 — Alíquota padrão'
        anexo = ''
        descricao_beneficio = ''
        if resultado.get('beneficios') and resultado['beneficios'].get('total_enquadramentos', 0) > 0:
            enqs = resultado['beneficios']['enquadramentos']
            if enqs:
                e = enqs[0]
                anexo = e.get('anexo', '')
                descricao_beneficio = e.get('descricao_anexo', '')
                base_legal = e.get('base_legal', f'LC 214/2025 — {anexo}')
        
        return {
            'cclasstrib_code': resultado.get('cclasstrib_code', '000001'),
            'cclasstrib_msg': resultado.get('cclasstrib_msg', '')[:120],
            'regime': resultado.get('regime', 'TRIBUTACAO_PADRAO'),
            'reducao_pct': resultado.get('reducao_pct', 0),
            'ibs_uf': resultado.get('ibs_uf', 0.10),
            'ibs_mun': resultado.get('ibs_mun', 0.025),
            'cbs': resultado.get('cbs', 0.90),
            'total_iva': resultado.get('total_iva', 1.025),
            'anexo': anexo,
            'descricao_beneficio': descricao_beneficio,
            'base_legal': base_legal,
            'requer_revisao_manual': requer_revisao,
            'opcoes_cclasstrib': opcoes_str
        }
    except Exception as e:
        logger.warning(f"Erro ao enriquecer NCM {ncm}: {e}")
        return fallback


def extract_zip_to_temp(zip_file) -> Tuple[str, List[str]]:
    """
    Extrai arquivo ZIP para diretório temporário e retorna lista de XMLs.
    
    Args:
        zip_file: Arquivo ZIP do Streamlit
        
    Returns:
        Tupla (diretório temporário, lista de caminhos de XMLs)
        
    Raises:
        Exception: Se houver erro ao extrair o ZIP
    """
    temp_dir = tempfile.mkdtemp(prefix="pricetax_batch_")
    xml_files = []
    
    try:
        with zipfile.ZipFile(zip_file, 'r') as zip_ref:
            # Extrair todos os arquivos
            zip_ref.extractall(temp_dir)
            
            # Buscar todos os XMLs (incluindo em subdiretórios)
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    if file.lower().endswith('.xml'):
                        xml_files.append(os.path.join(root, file))
    
    except zipfile.BadZipFile as e:
        logger.error(f"Arquivo ZIP corrompido ou inválido: {e}")
        raise Exception(f"Arquivo ZIP inválido: {str(e)}")
    except Exception as e:
        logger.error(f"Erro ao extrair ZIP: {e}")
        raise Exception(f"Erro ao extrair ZIP: {str(e)}")
    
    return temp_dir, xml_files


def validate_xml_structure(xml_path: str) -> Dict[str, Any]:
    """
    Valida estrutura básica do XML NFe com tratamento robusto de erros.
    
    Args:
        xml_path: Caminho para o arquivo XML
    
    Returns:
        Dict com status e mensagem de erro (se houver)
    """
    try:
        # Tentar fazer parse
        resultado = parse_nfe_xml(xml_path)
        
        # Verificar se houve erro no parse
        if 'erro' in resultado:
            logger.warning(f"Erro ao validar XML {xml_path}: {resultado.get('detalhes')}")
            return {'valid': False, 'error': resultado.get('detalhes', 'Erro desconhecido')}
        
        # Verificar campos obrigatórios
        if not resultado.get('emitente') or not resultado['emitente'].get('cnpj'):
            return {'valid': False, 'error': 'Emitente não encontrado ou CNPJ ausente'}
        
        if not resultado.get('destinatario'):
            return {'valid': False, 'error': 'Destinatário não encontrado'}
        
        if not resultado.get('itens') or len(resultado['itens']) == 0:
            return {'valid': False, 'error': 'Nenhum item encontrado na nota'}
        
        return {'valid': True, 'error': None}
    
    except Exception as e:
        logger.error(f"Erro inesperado ao validar XML {xml_path}: {e}")
        return {'valid': False, 'error': f'Erro inesperado: {str(e)}'}


def calculate_expected_ibs_cbs(item: Dict[str, Any]) -> Dict[str, float]:
    """
    Calcula IBS/CBS esperados com base líquida (2026) e alíquotas reais por NCM.
    
    Args:
        item: Dicionário com dados do item
        
    Returns:
        Dict com valores esperados
    """
    # Valores do item
    vprod = item.get('valor_total', 0.0)
    vfrete = item.get('vfrete', 0.0)
    vseg = item.get('vseg', 0.0)
    voutro = item.get('voutro', 0.0)
    vdesc = item.get('vdesc', 0.0)
    vicms = item.get('vicms', 0.0)
    vpis = item.get('vpis', 0.0)
    vcofins = item.get('vcofins', 0.0)
    
    # Base líquida (2026): vProd + vFrete + vSeg + vOutro - vDesc - ICMS - PIS - COFINS
    base_liquida = vprod + vfrete + vseg + voutro - vdesc - vicms - vpis - vcofins
    
    # Enriquecer com alíquotas reais por NCM+CFOP
    trib = item.get('_tributario', None)
    if trib:
        aliq_ibs = trib['ibs_uf'] + trib['ibs_mun']
        aliq_cbs = trib['cbs']
    else:
        aliq_ibs = 0.10 + 0.025  # 0,125% padrão
        aliq_cbs = 0.90  # 0,9%
    
    # Calcular valores esperados
    vibs_esperado = base_liquida * (aliq_ibs / 100)
    vcbs_esperado = base_liquida * (aliq_cbs / 100)
    
    return {
        'base_liquida': base_liquida,
        'vibs_esperado': vibs_esperado,
        'vcbs_esperado': vcbs_esperado,
        'aliq_ibs': aliq_ibs,
        'aliq_cbs': aliq_cbs
    }


def validate_ibs_cbs(item: Dict[str, Any], tolerancia: float = 0.02) -> Dict[str, Any]:
    """
    Valida se IBS/CBS do XML estão corretos.
    
    Args:
        item: Dicionário com dados do item
        tolerancia: Tolerância em reais para diferença
        
    Returns:
        Dict com status de validação
    """
    # Valores do XML
    vibs_xml = item.get('vibs', 0.0)
    vcbs_xml = item.get('vcbs', 0.0)
    
    # Calcular esperados
    esperados = calculate_expected_ibs_cbs(item)
    
    # Comparar
    diff_ibs = abs(vibs_xml - esperados['vibs_esperado'])
    diff_cbs = abs(vcbs_xml - esperados['vcbs_esperado'])
    
    ibs_ok = diff_ibs <= tolerancia
    cbs_ok = diff_cbs <= tolerancia
    
    return {
        'ibs_ok': ibs_ok,
        'cbs_ok': cbs_ok,
        'diff_ibs': diff_ibs,
        'diff_cbs': diff_cbs,
        'vibs_xml': vibs_xml,
        'vcbs_xml': vcbs_xml,
        'vibs_esperado': esperados['vibs_esperado'],
        'vcbs_esperado': esperados['vcbs_esperado'],
        'base_liquida': esperados['base_liquida']
    }


def process_single_xml(xml_path: str, save_to_collector: bool = True) -> Dict[str, Any]:
    """
    Processa um único XML com validação completa.
    
    Args:
        xml_path: Caminho do arquivo XML
        save_to_collector: Se True, salva no data_collector (espião)
        
    Returns:
        Dict com resultado do processamento
    """
    resultado = {
        'arquivo': os.path.basename(xml_path),
        'status': 'ERRO',
        'mensagem': '',
        'chave_acesso': '',
        'emitente_razao': '',
        'emitente_cnpj': '',
        'emitente_uf': '',
        'destinatario_razao': '',
        'destinatario_cnpj': '',
        'destinatario_uf': '',
        'data_emissao': '',
        'total_itens': 0,
        'itens_conformes': 0,
        'itens_divergentes': 0,
        'valor_total_nfe': 0.0,
        'validacoes': []
    }
    
    try:
        # Validar estrutura
        validacao = validate_xml_structure(xml_path)
        if not validacao['valid']:
            resultado['mensagem'] = validacao['error']
            return resultado
        
        # Parse completo
        dados = parse_nfe_xml(xml_path)
        
        # Extrair chave de acesso do nome do arquivo
        nome_arquivo = os.path.basename(xml_path)
        if len(nome_arquivo) >= 44:
            chave = nome_arquivo[:44]
            if chave.isdigit():
                resultado['chave_acesso'] = chave
        
        # Dados gerais
        resultado['emitente_razao'] = dados['emitente'].get('razao_social', '')
        resultado['emitente_cnpj'] = dados['emitente'].get('cnpj', '')
        resultado['emitente_uf'] = dados['emitente'].get('uf', '')
        resultado['destinatario_razao'] = dados['destinatario'].get('razao_social', '')
        resultado['destinatario_cnpj'] = dados['destinatario'].get('cnpj', '')
        resultado['destinatario_uf'] = dados['destinatario'].get('uf', '')
        resultado['data_emissao'] = dados.get('data_emissao', '')
        resultado['total_itens'] = len(dados['itens'])
        
        # Validar cada item
        itens_conformes = 0
        itens_divergentes = 0
        valor_total = 0.0
        
        for idx, item in enumerate(dados['itens'], 1):
            valor_total += item.get('valor_total', 0.0)
            
            # Enriquecer item com cClassTrib, alíquotas reais e base legal
            trib = enriquecer_item_tributario(
                ncm=item.get('ncm', ''),
                cfop=item.get('cfop', '')
            )
            item['_tributario'] = trib  # Injetar no item para uso em calculate_expected_ibs_cbs
            
            # Validar IBS/CBS com alíquotas reais
            validacao_item = validate_ibs_cbs(item)
            
            if validacao_item['ibs_ok'] and validacao_item['cbs_ok']:
                itens_conformes += 1
                status_item = 'CONFORME'
            else:
                itens_divergentes += 1
                status_item = 'DIVERGENTE'
            
            resultado['validacoes'].append({
                'item': idx,
                'ncm': item.get('ncm', ''),
                'cfop': item.get('cfop', ''),
                'descricao': item.get('descricao', ''),
                'valor': item.get('valor_total', 0.0),
                'status': status_item,
                'ibs_xml': validacao_item['vibs_xml'],
                'ibs_esperado': validacao_item['vibs_esperado'],
                'diff_ibs': validacao_item['diff_ibs'],
                'cbs_xml': validacao_item['vcbs_xml'],
                'cbs_esperado': validacao_item['vcbs_esperado'],
                'diff_cbs': validacao_item['diff_cbs'],
                'base_liquida': validacao_item['base_liquida'],
                # Campos tributários enriquecidos
                'cclasstrib': trib['cclasstrib_code'],
                'cclasstrib_msg': trib['cclasstrib_msg'],
                'regime': trib['regime'],
                'reducao_pct': trib['reducao_pct'],
                'ibs_uf_pct': trib['ibs_uf'],
                'ibs_mun_pct': trib['ibs_mun'],
                'cbs_pct': trib['cbs'],
                'total_iva_pct': trib['total_iva'],
                'anexo_lc214': trib['anexo'],
                'descricao_beneficio': trib['descricao_beneficio'],
                'base_legal': trib['base_legal'],
                'requer_revisao_manual': trib.get('requer_revisao_manual', False),
                'opcoes_cclasstrib': trib.get('opcoes_cclasstrib', '')
            })
        
        resultado['itens_conformes'] = itens_conformes
        resultado['itens_divergentes'] = itens_divergentes
        resultado['valor_total_nfe'] = valor_total
        
        # Status geral
        if itens_divergentes == 0:
            resultado['status'] = 'CONFORME'
            resultado['mensagem'] = f'Todos os {itens_conformes} itens estão corretos'
        else:
            resultado['status'] = 'DIVERGENTE'
            resultado['mensagem'] = f'{itens_divergentes} de {resultado["total_itens"]} itens com divergência'
        
        # Salvar no data_collector (espião)
        if save_to_collector:
            try:
                save_nfe_data(dados, usuario="BatchProcessor", anonimizar=False)
            except Exception as e:
                # Não falhar o processamento se o espião falhar
                print(f"Aviso: Erro ao salvar no data_collector: {e}")
        
    except Exception as e:
        resultado['mensagem'] = f'Erro: {str(e)}'
        resultado['status'] = 'ERRO'
    
    return resultado


def process_batch(xml_files: List[str], save_to_collector: bool = True, 
                  progress_callback=None) -> List[Dict[str, Any]]:
    """
    Processa lote de XMLs.
    
    Args:
        xml_files: Lista de caminhos de XMLs
        save_to_collector: Se True, salva no data_collector
        progress_callback: Função para atualizar progresso (opcional)
        
    Returns:
        Lista de resultados
    """
    resultados = []
    total = len(xml_files)
    
    for idx, xml_path in enumerate(xml_files, 1):
        resultado = process_single_xml(xml_path, save_to_collector)
        resultados.append(resultado)
        
        # Callback de progresso
        if progress_callback:
            progress_callback(idx, total, resultado['arquivo'])
    
    return resultados


def generate_summary_stats(resultados: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Gera estatísticas resumidas do processamento.
    
    Returns:
        Dict com estatísticas
    """
    total = len(resultados)
    conformes = sum(1 for r in resultados if r['status'] == 'CONFORME')
    divergentes = sum(1 for r in resultados if r['status'] == 'DIVERGENTE')
    erros = sum(1 for r in resultados if r['status'] == 'ERRO')
    
    total_itens = sum(r['total_itens'] for r in resultados)
    itens_conformes = sum(r['itens_conformes'] for r in resultados)
    itens_divergentes = sum(r['itens_divergentes'] for r in resultados)
    
    valor_total = sum(r['valor_total_nfe'] for r in resultados)
    
    return {
        'total_xmls': total,
        'xmls_conformes': conformes,
        'xmls_divergentes': divergentes,
        'xmls_erros': erros,
        'percentual_conformes': (conformes / total * 100) if total > 0 else 0,
        'total_itens': total_itens,
        'itens_conformes': itens_conformes,
        'itens_divergentes': itens_divergentes,
        'valor_total': valor_total,
        'data_processamento': datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    }


def _build_dados_completos_expandidos(resultados: List[Dict[str, Any]]) -> List[Dict]:
    """
    Constrói a lista de linhas para a aba 'Dados Completos' do Excel.

    LOGICA CENTRAL DE DUPLICACAO DE LINHAS:
    ----------------------------------------
    Para cada item de cada XML processado:

    a) NCM AMBIGUO (len(opcoes) > 1):
       Gera N linhas — uma por cClassTrib possível.
       Cada linha recebe:
         - 'ALERTA'     : mensagem de atencao com NCM e quantidade de opcoes
         - 'Tipo Linha' : 'OPCAO X DE N' (ex: 'OPCAO 1 DE 4')
         - 'cClassTrib' : codigo da opcao especifica desta linha
         - 'Descricao cClassTrib': descricao oficial da opcao
         - 'Situacao Operacao'  : criterio de distincao (ex: 'EAC acima do minimo')
         - 'Base Legal cClassTrib': artigo da LC 214/2025
       O usuario deve manter apenas a linha correta e excluir as demais.

    b) NCM UNICO (len(opcoes) <= 1):
       Gera 1 linha com:
         - 'ALERTA'     : 'TRATAMENTO UNICO: NCM possui classificacao tributaria definida.'
         - 'Tipo Linha' : 'UNICO'
         - cClassTrib e descricao do processamento normal

    NCMs ambíguos mapeados (cclasstrib_mapping.py):
        27101259 — Gasolina (4 opcoes: 620001, 620004, 620005, 620006)
        27101921 — Diesel   (2 opcoes: 620001, 620006)
        84212990 — Filtros medicos Anexo IV (2 opcoes: 200030, 200005)

    Args:
        resultados: Lista de dicts retornados por process_single_xml() ou process_batch().

    Returns:
        Lista de dicts prontos para pd.DataFrame(), com colunas 'ALERTA' e 'Tipo Linha'
        como primeiras colunas.
    """
    import re as _re
    # Importar com fallback seguro caso cclasstrib_mapping nao esteja disponivel
    try:
        from cclasstrib_mapping import get_opcoes_cclasstrib_por_ncm as _get_opts
    except ImportError:
        def _get_opts(ncm): return []

    linhas = []
    for r in resultados:
        for val in r['validacoes']:
            # Normalizar NCM: remover pontos, hifens e espacos
            ncm_clean = _re.sub(r'\D+', '', str(val.get('ncm', '')))
            # Consultar opcoes de cClassTrib para este NCM
            opcoes = _get_opts(ncm_clean)
            # NCM e ambiguo se houver mais de uma opcao de classificacao tributaria
            ambiguo = len(opcoes) > 1

            base = {
                'Arquivo': r['arquivo'],
                'Chave Acesso': r['chave_acesso'],
                'Emitente': r['emitente_razao'],
                'Destinatário': r['destinatario_razao'],
                'Item': val['item'],
                'NCM': val['ncm'],
                'CFOP': val['cfop'],
                'Descrição': val['descricao'],
                'Valor Item (R$)': val['valor'],
                'Base Líquida (R$)': val['base_liquida'],
                'Regime IVA': val.get('regime', ''),
                'Redução (%)': val.get('reducao_pct', 0),
                'IBS UF (%)': val.get('ibs_uf_pct', 0.10),
                'IBS Municipal (%)': val.get('ibs_mun_pct', 0.025),
                'CBS (%)': val.get('cbs_pct', 0.90),
                'Total IVA (%)': val.get('total_iva_pct', 1.025),
                'Anexo LC 214/2025': val.get('anexo_lc214', ''),
                'Benefício Fiscal': val.get('descricao_beneficio', ''),
                'Base Legal': val.get('base_legal', 'LC 214/2025'),
                'IBS XML (R$)': val['ibs_xml'],
                'IBS Esperado (R$)': val['ibs_esperado'],
                'Dif. IBS (R$)': val['diff_ibs'],
                'CBS XML (R$)': val['cbs_xml'],
                'CBS Esperado (R$)': val['cbs_esperado'],
                'Dif. CBS (R$)': val['diff_cbs'],
                'Status Validação': val['status'],
            }

            if ambiguo:
                # Uma linha por opcao de cClassTrib
                for i, op in enumerate(opcoes, 1):
                    linha = dict(base)
                    linha['ALERTA'] = f'ATENCAO: NCM {ncm_clean} possui {len(opcoes)} cClassTribs possiveis. Selecione a opcao correta para a sua operacao.'
                    linha['Tipo Linha'] = f'OPCAO {i} DE {len(opcoes)}'
                    linha['cClassTrib'] = op['code']
                    linha['Descricao cClassTrib'] = op['descricao']
                    linha['Situacao Operacao'] = op['situacao']
                    linha['Base Legal cClassTrib'] = op['base_legal']
                    linhas.append(linha)
            else:
                # Linha unica com alerta de tratamento unico
                linha = dict(base)
                linha['ALERTA'] = 'TRATAMENTO UNICO: NCM possui classificacao tributaria definida.'
                linha['Tipo Linha'] = 'UNICO'
                linha['cClassTrib'] = val.get('cclasstrib', '')
                linha['Descricao cClassTrib'] = val.get('cclasstrib_msg', '')
                linha['Situacao Operacao'] = ''
                linha['Base Legal cClassTrib'] = val.get('base_legal', 'LC 214/2025')
                linhas.append(linha)

    # Reordenar colunas: ALERTA e Tipo Linha primeiro
    col_order = [
        'ALERTA', 'Tipo Linha', 'Arquivo', 'Chave Acesso', 'Emitente', 'Destinatario',
        'Item', 'NCM', 'CFOP', 'Descricao', 'Valor Item (R$)', 'Base Liquida (R$)',
        'cClassTrib', 'Descricao cClassTrib', 'Situacao Operacao', 'Base Legal cClassTrib',
        'Regime IVA', 'Reducao (%)', 'IBS UF (%)', 'IBS Municipal (%)', 'CBS (%)',
        'Total IVA (%)', 'Anexo LC 214/2025', 'Beneficio Fiscal', 'Base Legal',
        'IBS XML (R$)', 'IBS Esperado (R$)', 'Dif. IBS (R$)',
        'CBS XML (R$)', 'CBS Esperado (R$)', 'Dif. CBS (R$)', 'Status Validacao'
    ]
    # Usar as chaves reais (com acento) que foram definidas no dict base
    return linhas


def _aplicar_formatacao_excel(writer, sheet_name: str, df: pd.DataFrame,
                               col_alerta: str = 'ALERTA',
                               col_tipo: str = 'Tipo Linha') -> None:
    """
    Aplica formatacao visual profissional (openpyxl) na aba 'Dados Completos'.

    Paleta de cores PRICETAX:
        Cabecalho   : fundo #1A1A1A (preto) + texto #FFDD00 (amarelo PRICETAX)
        NCM ambiguo : fundo #FFFFF3CD (laranja claro) + texto #7B3F00 (marrom)
                      Coluna ALERTA: fundo #FFC107 (ambar) + texto negrito
        NCM unico   : fundo #E8F5E9 (verde claro) + texto #1B5E20 (verde escuro)
                      Coluna ALERTA: fundo #4CAF50 (verde) + texto negrito

    Recursos aplicados:
        - Cabecalho fixo (freeze_panes = 'A2')
        - Filtro automatico em todas as colunas
        - Quebra de texto (wrap_text=True) para legibilidade
        - Borda fina em todas as celulas
        - Largura de coluna ajustada automaticamente (max 60 caracteres)

    Args:
        writer     : pd.ExcelWriter com engine='openpyxl' ainda aberto.
        sheet_name : Nome da aba a formatar (deve existir em writer.sheets).
        df         : DataFrame que foi escrito na aba (usado para calcular larguras).
        col_alerta : Nome da coluna de alertas (padrao: 'ALERTA').
        col_tipo   : Nome da coluna de tipo de linha (padrao: 'Tipo Linha').
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ws = writer.sheets[sheet_name]

    # Cores
    COR_HEADER_BG    = 'FF1A1A1A'  # preto
    COR_HEADER_FG    = 'FFFFDD00'  # amarelo PRICETAX
    COR_AMBIGUO_BG   = 'FFFFF3CD'  # laranja muito claro (aviso)
    COR_AMBIGUO_FG   = 'FF7B3F00'  # marrom escuro
    COR_UNICO_BG     = 'FFE8F5E9'  # verde muito claro
    COR_UNICO_FG     = 'FF1B5E20'  # verde escuro
    COR_ALERTA_BG    = 'FFFFC107'  # amarelo ambar (alerta ambiguo)
    COR_ALERTA_UNICO = 'FF4CAF50'  # verde (alerta unico)

    fill_header    = PatternFill('solid', fgColor=COR_HEADER_BG)
    font_header    = Font(bold=True, color=COR_HEADER_FG, size=10)
    fill_ambiguo   = PatternFill('solid', fgColor=COR_AMBIGUO_BG)
    font_ambiguo   = Font(color=COR_AMBIGUO_FG, size=9)
    fill_unico     = PatternFill('solid', fgColor=COR_UNICO_BG)
    font_unico     = Font(color=COR_UNICO_FG, size=9)
    fill_alerta_a  = PatternFill('solid', fgColor=COR_ALERTA_BG)
    fill_alerta_u  = PatternFill('solid', fgColor=COR_ALERTA_UNICO)
    font_alerta    = Font(bold=True, size=9)
    align_wrap     = Alignment(wrap_text=True, vertical='top')
    thin_border    = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Identificar indices das colunas ALERTA e Tipo Linha
    cols = list(df.columns)
    idx_alerta = cols.index(col_alerta) + 1 if col_alerta in cols else None
    idx_tipo   = cols.index(col_tipo) + 1 if col_tipo in cols else None

    # Formatar cabecalho (linha 1)
    for cell in ws[1]:
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # Formatar linhas de dados
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        tipo_val = ''
        if idx_tipo:
            tipo_val = str(ws.cell(row=row_idx, column=idx_tipo).value or '')

        is_ambiguo = 'OPCAO' in tipo_val
        fill_row  = fill_ambiguo if is_ambiguo else fill_unico
        font_row  = font_ambiguo if is_ambiguo else font_unico

        for cell in row:
            cell.fill = fill_row
            cell.font = font_row
            cell.alignment = align_wrap
            cell.border = thin_border

        # Coluna ALERTA: formatacao especial
        if idx_alerta:
            cell_alerta = ws.cell(row=row_idx, column=idx_alerta)
            cell_alerta.fill = fill_alerta_a if is_ambiguo else fill_alerta_u
            cell_alerta.font = Font(bold=True, size=9,
                                    color='FF7B3F00' if is_ambiguo else 'FF1B5E20')

    # Ajustar largura das colunas
    for col_idx, col in enumerate(df.columns, 1):
        max_len = max(
            len(str(col)),
            df[col].astype(str).str.len().max() if len(df) > 0 else 0
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)

    # Congelar cabecalho
    ws.freeze_panes = 'A2'

    # Filtro automatico
    ws.auto_filter.ref = ws.dimensions


def generate_excel_report(resultados: List[Dict[str, Any]]) -> BytesIO:
    """
    Gera relatorio Excel profissional com 4 abas.
    - Aba 'Dados Completos': NCMs ambiguos duplicam linhas (uma por cClassTrib possivel)
      com formatacao visual diferenciada e alerta na coluna ALERTA.
    
    Returns:
        BytesIO com arquivo Excel
    """
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # ----------------------------------------------------------------
        # ABA 1: RESUMO
        # ----------------------------------------------------------------
        stats = generate_summary_stats(resultados)

        # Contar NCMs ambiguos no lote
        import re as _re
        try:
            from cclasstrib_mapping import get_opcoes_cclasstrib_por_ncm as _get_opts
        except ImportError:
            def _get_opts(ncm): return []

        ncms_ambiguos = set()
        itens_ambiguos = 0
        for r in resultados:
            for val in r.get('validacoes', []):
                ncm_c = _re.sub(r'\D+', '', str(val.get('ncm', '')))
                if len(_get_opts(ncm_c)) > 1:
                    ncms_ambiguos.add(ncm_c)
                    itens_ambiguos += 1

        df_resumo = pd.DataFrame([
            ['Total de XMLs Processados', stats['total_xmls']],
            ['XMLs Conformes', stats['xmls_conformes']],
            ['XMLs com Divergencias', stats['xmls_divergentes']],
            ['XMLs com Erros', stats['xmls_erros']],
            ['% Conformidade', f"{stats['percentual_conformes']:.2f}%"],
            ['', ''],
            ['Total de Itens', stats['total_itens']],
            ['Itens Conformes', stats['itens_conformes']],
            ['Itens Divergentes', stats['itens_divergentes']],
            ['', ''],
            ['NCMs com Multiplos cClassTribs (Revisao Manual)', len(ncms_ambiguos)],
            ['Itens que Requerem Revisao Manual', itens_ambiguos],
            ['NCMs Ambiguos Identificados', ', '.join(sorted(ncms_ambiguos)) if ncms_ambiguos else 'Nenhum'],
            ['', ''],
            ['Valor Total (R$)', f"{stats['valor_total']:,.2f}"],
            ['', ''],
            ['Data do Processamento', stats['data_processamento']]
        ], columns=['Metrica', 'Valor'])
        df_resumo.to_excel(writer, sheet_name='Resumo', index=False)

        # Formatar aba Resumo
        from openpyxl.styles import PatternFill, Font, Alignment
        ws_res = writer.sheets['Resumo']
        for cell in ws_res[1]:
            cell.fill = PatternFill('solid', fgColor='FF1A1A1A')
            cell.font = Font(bold=True, color='FFFFDD00', size=10)
        # Destacar linhas de NCMs ambiguos
        for row in ws_res.iter_rows(min_row=2):
            val_cell = row[0].value or ''
            if 'Ambiguo' in str(val_cell) or 'Revisao' in str(val_cell):
                for cell in row:
                    cell.fill = PatternFill('solid', fgColor='FFFFF3CD')
                    cell.font = Font(bold=True, color='FF7B3F00', size=10)
        ws_res.column_dimensions['A'].width = 50
        ws_res.column_dimensions['B'].width = 40
        ws_res.freeze_panes = 'A2'

        # ----------------------------------------------------------------
        # ABA 2: VALIDACAO (todos os XMLs)
        # ----------------------------------------------------------------
        df_validacao = pd.DataFrame([{
            'Arquivo': r['arquivo'],
            'Status': r['status'],
            'Chave Acesso': r['chave_acesso'],
            'Emitente': r['emitente_razao'],
            'CNPJ Emitente': r['emitente_cnpj'],
            'UF Emitente': r['emitente_uf'],
            'Destinatario': r['destinatario_razao'],
            'CNPJ Destinatario': r['destinatario_cnpj'],
            'UF Destinatario': r['destinatario_uf'],
            'Data Emissao': r['data_emissao'],
            'Total Itens': r['total_itens'],
            'Itens Conformes': r['itens_conformes'],
            'Itens Divergentes': r['itens_divergentes'],
            'Valor Total (R$)': r['valor_total_nfe'],
            'Mensagem': r['mensagem']
        } for r in resultados])
        df_validacao.to_excel(writer, sheet_name='Validacao', index=False)

        ws_val = writer.sheets['Validacao']
        for cell in ws_val[1]:
            cell.fill = PatternFill('solid', fgColor='FF1A1A1A')
            cell.font = Font(bold=True, color='FFFFDD00', size=10)
        for row in ws_val.iter_rows(min_row=2):
            status = str(row[1].value or '')
            if status == 'DIVERGENTE':
                for cell in row:
                    cell.fill = PatternFill('solid', fgColor='FFFFF3CD')
            elif status == 'ERRO':
                for cell in row:
                    cell.fill = PatternFill('solid', fgColor='FFFCE4EC')
            elif status == 'CONFORME':
                for cell in row:
                    cell.fill = PatternFill('solid', fgColor='FFE8F5E9')
        ws_val.freeze_panes = 'A2'
        ws_val.auto_filter.ref = ws_val.dimensions

        # ----------------------------------------------------------------
        # ABA 3: DIVERGENCIAS
        # ----------------------------------------------------------------
        divergentes_list = [r for r in resultados if r['status'] in ['DIVERGENTE', 'ERRO']]
        if divergentes_list:
            df_divergencias = pd.DataFrame([{
                'Arquivo': r['arquivo'],
                'Status': r['status'],
                'Chave Acesso': r['chave_acesso'],
                'Emitente': r['emitente_razao'],
                'Destinatario': r['destinatario_razao'],
                'Itens Divergentes': r['itens_divergentes'],
                'Valor Total (R$)': r['valor_total_nfe'],
                'Mensagem': r['mensagem']
            } for r in divergentes_list])
            df_divergencias.to_excel(writer, sheet_name='Divergencias', index=False)
            ws_div = writer.sheets['Divergencias']
            for cell in ws_div[1]:
                cell.fill = PatternFill('solid', fgColor='FF1A1A1A')
                cell.font = Font(bold=True, color='FFFFDD00', size=10)
            ws_div.freeze_panes = 'A2'
            ws_div.auto_filter.ref = ws_div.dimensions

        # ----------------------------------------------------------------
        # ABA 4: DADOS COMPLETOS — com duplicacao de linhas para NCMs ambiguos
        # ----------------------------------------------------------------
        linhas_expandidas = _build_dados_completos_expandidos(resultados)

        if linhas_expandidas:
            df_completo = pd.DataFrame(linhas_expandidas)

            # Garantir que colunas ALERTA e Tipo Linha sejam as primeiras
            cols_priority = ['ALERTA', 'Tipo Linha']
            other_cols = [c for c in df_completo.columns if c not in cols_priority]
            df_completo = df_completo[cols_priority + other_cols]

            df_completo.to_excel(writer, sheet_name='Dados Completos', index=False)
            _aplicar_formatacao_excel(writer, 'Dados Completos', df_completo,
                                      col_alerta='ALERTA', col_tipo='Tipo Linha')
    
    output.seek(0)
    return output

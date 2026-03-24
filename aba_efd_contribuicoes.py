"""
aba_efd_contribuicoes.py — Módulo EFD Contribuições (SPED PIS/COFINS)
=======================================================================

Processa arquivos EFD Contribuições (.txt) e deriva automaticamente
o cClassTrib IBS/CBS por NCM.

ESTRUTURA REAL DO ARQUIVO (confirmada por análise empírica):
------------------------------------------------------------

Bloco C — Documentos Fiscais:
  C100: cabeçalho NF-e/NFC-e (IND_OPER=1 para saída)
  C175: itens de NFC-e (modelo 65 — varejo/cupom fiscal)
        [2]=CFOP [3]=VL_ITEM [4]=VL_DESC [5]=CST_PIS [6]=VL_BC_PIS [7]=ALIQ_PIS
        [8]=VL_PIS [9]=COD_CTA [10]=VL_COFINS [11]=CST_COFINS [12]=VL_BC_COFINS
        [13]=ALIQ_COFINS [14]=VL_COFINS2 [15]=COD_CTA2 [16]=VL_COFINS3 [17]=COD_MUN_ORIG
  C180: consolidado de NF-e modelo 55 por item/período
        [2]=COD_MOD [3]=DT_INI [4]=DT_FIN [5]=COD_ITEM [6]=NCM [7]=EX_IPI [8]=VL_TOTAL
  C181: detalhe PIS do C180 (campos INVERTIDOS neste arquivo — bug do emissor)
        Campos esperados: [3]=CST_PIS [4]=CFOP
        Campos reais:     [3]=CFOP    [4]=VL_ITEM (campos trocados)
  C185: detalhe COFINS do C180 (mesma inversão)

Bloco M — Apuração:
  M200: apuração PIS não cumulativo
  M600: apuração COFINS não cumulativo

Bloco 0 — Abertura:
  0000: identificação da empresa
  0111: receita bruta por tipo
  0200: cadastro de produtos (COD_ITEM → NCM + descrição)

MODO DE PROCESSAMENTO:
  - C175 (NFC-e): sem NCM individual → agrupado por CFOP, sem cClassTrib por NCM
  - C180 (NF-e modelo 55): COM NCM → derivação cClassTrib via BDBENEF + cclasstrib_mapping
  - Ambos podem coexistir no mesmo arquivo (como neste caso)

Anti-regressão: este arquivo é ISOLADO do app.py.
Não modifica nenhuma função existente. Apenas adiciona nova aba.

Autor: PRICETAX
Data: 24/03/2026
"""

import io
import re
from collections import defaultdict
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st

# =============================================================================
# CONSTANTES DE ALÍQUOTA — ANO TESTE 2026
# =============================================================================

IBS_UF_REF  = 0.10    # 0,10% IBS Estadual
IBS_MUN_REF = 0.025   # 0,025% IBS Municipal
CBS_REF     = 0.90    # 0,90% CBS
TOTAL_IVA_REF = IBS_UF_REF + IBS_MUN_REF + CBS_REF  # 1,025%

# =============================================================================
# BASE LEGAL POR REGIME
# =============================================================================

BASE_LEGAL_MAP = {
    "TRIBUTACAO_PADRAO":               "LC 214/2025, Art. 20 — Tributação integral (alíquota cheia)",
    "ALIQ_ZERO_CESTA_BASICA":          "LC 214/2025, Anexo I — Cesta Básica Nacional (alíquota zero)",
    "RED_60_ESSENCIALIDADE":           "LC 214/2025, Art. 21, § 1º — Redução de 60%",
    "RED_60":                          "LC 214/2025, Art. 21, § 1º — Redução de 60%",
    "RED_30":                          "LC 214/2025, Art. 21, § 2º — Redução de 30% (Cesta Estendida)",
    "MONOFASICO_COMBUSTIVEL":          "LC 214/2025, Art. 172 — Tributação Monofásica (combustíveis)",
    "INSUMO_AGROPECUARIO_ANEXO_IX":    "LC 214/2025, Anexo IX — Insumos Agropecuários",
    "DIFERIMENTO_INSUMO_AGROPECUARIO": "LC 214/2025, Art. 138 — Diferimento (insumos agropecuários)",
}

# =============================================================================
# HELPERS
# =============================================================================

def _normalizar_ncm(ncm_raw: str) -> str:
    """Remove caracteres não numéricos e padeia com zeros à esquerda (8 dígitos)."""
    return re.sub(r"\D", "", ncm_raw).zfill(8)


def _parse_decimal(valor: str) -> float:
    """Converte string decimal brasileira (vírgula) para float."""
    try:
        return float(valor.strip().replace(".", "").replace(",", "."))
    except (ValueError, AttributeError):
        return 0.0


def _detectar_c181_invertido(campos: List[str]) -> bool:
    """
    Detecta se o C181/C185 está com campos invertidos (bug de alguns emissores).

    Sintoma confirmado neste arquivo:
      - campo[3] contém CFOP (4 dígitos, ex: '5929') em vez de CST (2-3 dígitos)
      - campo[4] contém VL_ITEM (valor monetário) em vez de CFOP

    Retorna True se os campos estão invertidos.
    """
    if len(campos) < 5:
        return False
    cst_candidato  = campos[3].strip() if len(campos) > 3 else ""
    cfop_candidato = campos[4].strip() if len(campos) > 4 else ""
    # CST válido: 2-3 dígitos numéricos (ex: '01', '06', '49')
    # CFOP: exatamente 4 dígitos (ex: '5929', '5102')
    if re.match(r"^\d{4}$", cst_candidato):
        return True
    # Se o campo[4] parece valor monetário → invertido
    if re.match(r"^\d+[,\.]\d+$", cfop_candidato):
        return True
    return False


# =============================================================================
# PARSER PRINCIPAL
# =============================================================================

def processar_efd_contribuicoes(conteudo: str) -> Dict[str, Any]:
    """
    Processa o conteúdo de um arquivo EFD Contribuições (.txt).

    Suporta arquivos com C175 (NFC-e) e/ou C180 (NF-e modelo 55) — ambos
    podem coexistir no mesmo arquivo.

    Retorna dicionário com:
      - modo: "C175_NFCE" | "C180_NFE" | "MISTO" | "DESCONHECIDO"
      - empresa: dict com CNPJ, nome, período
      - itens_c180: lista de itens C180 (com NCM real → derivação cClassTrib)
      - itens_c175: lista de itens C175 agrupados por CFOP (sem NCM individual)
      - apuracao_pis: dict com valores M200
      - apuracao_cofins: dict com valores M600
      - receita_bruta: dict com valores 0111
      - total_notas_nfce: int (contagem de NFC-e no C100)
      - alertas: lista de avisos
    """
    linhas = conteudo.splitlines()

    resultado = {
        "modo":            "DESCONHECIDO",
        "empresa":         {},
        "itens_c180":      [],   # NF-e modelo 55 — com NCM real
        "itens_c175":      [],   # NFC-e — agrupado por CFOP
        "apuracao_pis":    {},
        "apuracao_cofins": {},
        "receita_bruta":   {},
        "total_notas_nfce": 0,
        "alertas":         [],
    }

    # ── Fase 1: Mapa de produtos (0200) e empresa (0000 / 0111) ──────────────
    prod_map: Dict[str, Dict[str, str]] = {}  # COD_ITEM → {ncm, descricao}

    for linha in linhas:
        campos = linha.split("|")
        if len(campos) < 3:
            continue
        reg = campos[1].strip()

        if reg == "0000" and len(campos) > 9:
            # Estrutura real do 0000 (EFD Contribuições):
            # [2]=VERSAO [3]=TIPO_ESCRIT [4]=IND_SIT_ESP [5]=NUM_REC_ANTERIOR
            # [6]=DT_INI [7]=DT_FIN [8]=NOME [9]=CNPJ [10]=UF
            resultado["empresa"] = {
                "cnpj":        campos[9].strip() if len(campos) > 9 else "",
                "nome":        campos[8].strip() if len(campos) > 8 else "",
                "periodo":     _fmt_data(campos[6].strip()) if len(campos) > 6 else "",
                "periodo_fim": _fmt_data(campos[7].strip()) if len(campos) > 7 else "",
                "uf":          campos[10].strip() if len(campos) > 10 else "",
                "versao":      campos[2].strip() if len(campos) > 2 else "",
            }

        elif reg == "0111" and len(campos) > 5:
            resultado["receita_bruta"] = {
                "tributada_cumulativo":     _parse_decimal(campos[2]) if len(campos) > 2 else 0.0,
                "tributada_nao_cumulativo": _parse_decimal(campos[3]) if len(campos) > 3 else 0.0,
                "nao_tributada":            _parse_decimal(campos[4]) if len(campos) > 4 else 0.0,
                "exportacao":               _parse_decimal(campos[5]) if len(campos) > 5 else 0.0,
            }

        elif reg == "0200" and len(campos) > 8:
            cod_item = campos[2].strip()
            descr    = campos[3].strip()
            ncm_raw  = campos[8].strip()
            ncm      = _normalizar_ncm(ncm_raw)
            if cod_item:
                prod_map[cod_item] = {"ncm": ncm, "descricao": descr}

    # ── Fase 2: Processar C175 (NFC-e) ───────────────────────────────────────
    # C175: [2]=CFOP [3]=VL_ITEM [4]=VL_DESC [5]=CST_PIS
    c175_por_cfop: Dict[str, Dict] = defaultdict(lambda: {"vl_total": 0.0, "qtd": 0, "cst_pis": ""})
    cfop_saida = re.compile(r"^[567]\d{3}$")

    current_ind_oper = ""
    current_cod_mod  = ""

    for linha in linhas:
        campos = linha.split("|")
        if len(campos) < 3:
            continue
        reg = campos[1].strip()

        if reg == "C100" and len(campos) > 5:
            current_ind_oper = campos[2].strip()
            current_cod_mod  = campos[5].strip()  # 55=NF-e, 65=NFC-e
            if current_ind_oper == "1":
                resultado["total_notas_nfce"] += 1

        elif reg == "C175" and current_ind_oper == "1":
            cfop    = campos[2].strip() if len(campos) > 2 else "5102"
            vl_item = _parse_decimal(campos[3]) if len(campos) > 3 else 0.0
            cst_pis = campos[5].strip() if len(campos) > 5 else ""
            if not cfop:
                cfop = "5102"
            if cfop_saida.match(cfop) and vl_item > 0:
                c175_por_cfop[cfop]["vl_total"] += vl_item
                c175_por_cfop[cfop]["qtd"]      += 1
                c175_por_cfop[cfop]["cst_pis"]   = cst_pis  # último CST visto

        elif reg in ("C190", "C300", "D100", "E100"):
            current_ind_oper = ""
            current_cod_mod  = ""

    # Converter C175 agrupado para lista
    for cfop, dados in c175_por_cfop.items():
        resultado["itens_c175"].append({
            "cfop":      cfop,
            "descricao": f"NFC-e — CFOP {cfop}",
            "ncm":       "N/A",
            "cst_pis":   dados["cst_pis"],
            "vl_total":  dados["vl_total"],
            "qtd_notas": dados["qtd"],
            "origem":    "C175_NFCE",
        })

    # ── Fase 3: Processar C180 (NF-e modelo 55) ──────────────────────────────
    # C180: [2]=COD_MOD [3]=DT_INI [4]=DT_FIN [5]=COD_ITEM [6]=NCM [7]=EX_IPI [8]=VL_TOTAL
    c180_atual: Optional[Dict] = None
    c181_invertido_detectado   = False

    for linha in linhas:
        campos = linha.split("|")
        if len(campos) < 3:
            continue
        reg = campos[1].strip()

        if reg == "C180":
            # Flush do C180 anterior sem C181/C185
            if c180_atual is not None:
                resultado["itens_c180"].append(dict(c180_atual))

            cod_item = campos[5].strip() if len(campos) > 5 else ""
            ncm_raw  = campos[6].strip() if len(campos) > 6 else ""
            ncm      = _normalizar_ncm(ncm_raw)
            vl_total = _parse_decimal(campos[8]) if len(campos) > 8 else 0.0
            dt_ini   = _fmt_data(campos[3].strip()) if len(campos) > 3 else ""
            dt_fin   = _fmt_data(campos[4].strip()) if len(campos) > 4 else ""

            # Enriquecer descrição pelo 0200
            descricao = prod_map.get(cod_item, {}).get("descricao", cod_item)

            c180_atual = {
                "ncm":      ncm,
                "cod_item": cod_item,
                "descricao": descricao,
                "cfop":     "5102",  # padrão; será atualizado pelo C181
                "cst_pis":  "",
                "vl_total": vl_total,
                "dt_ini":   dt_ini,
                "dt_fin":   dt_fin,
                "origem":   "C180_NFE55",
            }

        elif reg in ("C181", "C185") and c180_atual is not None:
            # Detectar inversão de campos (bug do emissor)
            invertido = _detectar_c181_invertido(campos)

            if invertido and not c181_invertido_detectado:
                c181_invertido_detectado = True
                resultado["alertas"].append(
                    "C181/C185 com campos invertidos detectado (bug do emissor): "
                    "campos CST e CFOP estão trocados. Correção automática aplicada."
                )

            if invertido:
                # Estrutura real quando invertido (confirmada empiricamente):
                # [2]=CST_PIS [3]=CFOP [4]=VL_ITEM [5]=VL_BC_PIS [6]=ALIQ_PIS [7]=VL_PIS
                c180_atual["cst_pis"] = campos[2].strip() if len(campos) > 2 else ""
                c180_atual["cfop"]    = campos[3].strip() if len(campos) > 3 else "5102"
            else:
                # Campos normais: [3]=CST_PIS, [4]=CFOP
                c180_atual["cst_pis"] = campos[3].strip() if len(campos) > 3 else ""
                c180_atual["cfop"]    = campos[4].strip() if len(campos) > 4 else "5102"

            # Salvar apenas no C181 (não duplicar no C185)
            if reg == "C181":
                resultado["itens_c180"].append(dict(c180_atual))
                c180_atual = None

    # Flush final
    if c180_atual is not None:
        resultado["itens_c180"].append(dict(c180_atual))

    # ── Fase 4: Apuração PIS (M200) e COFINS (M600) ──────────────────────────
    for linha in linhas:
        campos = linha.split("|")
        if len(campos) < 3:
            continue
        reg = campos[1].strip()

        if reg == "M200" and len(campos) > 11:
            resultado["apuracao_pis"] = {
                "vl_tot_cont_nc_per":  _parse_decimal(campos[2]),
                "vl_cred_desc_nc_per": _parse_decimal(campos[3]),
                "vl_cred_desc_ant":    _parse_decimal(campos[4]),
                "vl_tot_cont_cum_per": _parse_decimal(campos[5]),
                "vl_ret_nc":           _parse_decimal(campos[6]),
                "vl_ret_cum":          _parse_decimal(campos[7]),
                "vl_out_ded_nc":       _parse_decimal(campos[8]),
                "vl_cont_nc_rec":      _parse_decimal(campos[9]),
                "vl_cont_cum_rec":     _parse_decimal(campos[10]),
                "vl_tot_cont_rec":     _parse_decimal(campos[11]),
            }

        elif reg == "M600" and len(campos) > 11:
            resultado["apuracao_cofins"] = {
                "vl_tot_cont_nc_per":  _parse_decimal(campos[2]),
                "vl_cred_desc_nc_per": _parse_decimal(campos[3]),
                "vl_cred_desc_ant":    _parse_decimal(campos[4]),
                "vl_tot_cont_cum_per": _parse_decimal(campos[5]),
                "vl_ret_nc":           _parse_decimal(campos[6]),
                "vl_ret_cum":          _parse_decimal(campos[7]),
                "vl_out_ded_nc":       _parse_decimal(campos[8]),
                "vl_cont_nc_rec":      _parse_decimal(campos[9]),
                "vl_cont_cum_rec":     _parse_decimal(campos[10]),
                "vl_tot_cont_rec":     _parse_decimal(campos[11]),
            }

    # ── Fase 5: Determinar modo ───────────────────────────────────────────────
    tem_c180 = len(resultado["itens_c180"]) > 0
    tem_c175 = len(resultado["itens_c175"]) > 0

    if tem_c180 and tem_c175:
        resultado["modo"] = "MISTO"
    elif tem_c180:
        resultado["modo"] = "C180_NFE55"
    elif tem_c175:
        resultado["modo"] = "C175_NFCE"
    else:
        resultado["alertas"].append("Nenhum item de saída encontrado (C175/C180). Verifique o arquivo.")

    return resultado


def _fmt_data(data_raw: str) -> str:
    """Formata DDMMAAAA para DD/MM/AAAA."""
    d = re.sub(r"\D", "", data_raw)
    if len(d) == 8:
        return f"{d[:2]}/{d[2:4]}/{d[4:]}"
    return data_raw


# =============================================================================
# ENRIQUECIMENTO cClassTrib POR NCM (apenas para itens C180 com NCM real)
# =============================================================================

def enriquecer_cclasstrib(itens: List[Dict], beneficios_engine, consulta_ncm_fn) -> List[Dict]:
    """
    Enriquece cada item C180 com cClassTrib, alíquotas IBS/CBS e base legal.

    Lógica de priorização:
    1. NCM com múltiplos cClassTribs (cclasstrib_mapping) → flag_ambiguo = True
    2. NCM com benefício fiscal (BDBENEF) → cClassTrib pelo Anexo
    3. Fallback → 000001 (tributação integral)

    Itens com NCM='N/A' (C175) são ignorados neste enriquecimento.
    """
    try:
        from cclasstrib_mapping import get_opcoes_cclasstrib_por_ncm, get_cclasstrib_by_anexo
    except ImportError:
        st.error("Módulo cclasstrib_mapping não encontrado.")
        return itens

    enriquecidos = []

    for item in itens:
        ncm  = item.get("ncm", "00000000")
        cfop = item.get("cfop", "5102")

        # Valores padrão
        cclasstrib       = "000001"
        descr_cclasstrib = "Tributação integral IBS/CBS"
        regime           = "TRIBUTACAO_PADRAO"
        reducao_pct      = 0
        anexo_beneficio  = ""
        descr_beneficio  = "Sem benefício fiscal identificado"
        flag_ambiguo     = False
        opcoes_str       = ""
        ibs_uf           = IBS_UF_REF
        ibs_mun          = IBS_MUN_REF
        cbs              = CBS_REF

        # Apenas derivar cClassTrib para NCMs reais (não N/A e não 00000000)
        if ncm and ncm not in ("N/A", "00000000"):
            # 1. Verificar NCMs com múltiplos cClassTribs
            try:
                opcoes_multi = get_opcoes_cclasstrib_por_ncm(ncm)
                if len(opcoes_multi) > 1:
                    flag_ambiguo     = True
                    cclasstrib       = opcoes_multi[0]["code"]
                    descr_cclasstrib = opcoes_multi[0].get("descricao", "")
                    opcoes_str       = " | ".join(
                        [f"{op['code']}: {op.get('situacao', '')}" for op in opcoes_multi]
                    )
                elif len(opcoes_multi) == 1:
                    cclasstrib       = opcoes_multi[0]["code"]
                    descr_cclasstrib = opcoes_multi[0].get("descricao", "")
            except Exception:
                pass

            # 2. Consultar BDBENEF (apenas se não ambíguo)
            if not flag_ambiguo and beneficios_engine and consulta_ncm_fn:
                try:
                    beneficios = consulta_ncm_fn(beneficios_engine, ncm)
                    if beneficios.get("total_enquadramentos", 0) > 0:
                        enq             = beneficios["enquadramentos"][0]
                        reducao         = enq.get("reducao_aliquota", 0)
                        anexo_beneficio = enq.get("anexo", "")
                        descr_beneficio = enq.get("descricao_anexo", "")
                        reducao_pct     = reducao

                        fator   = (100 - reducao) / 100
                        ibs_uf  = round(IBS_UF_REF  * fator, 6)
                        ibs_mun = round(IBS_MUN_REF * fator, 6)
                        cbs     = round(CBS_REF     * fator, 6)

                        if reducao == 100:
                            regime = "ALIQ_ZERO_CESTA_BASICA"
                        elif reducao == 60:
                            regime = f"RED_60_{anexo_beneficio}" if anexo_beneficio else "RED_60_ESSENCIALIDADE"
                        else:
                            regime = f"RED_{int(reducao)}"

                        code_res, msg_res = get_cclasstrib_by_anexo(reducao, anexo_beneficio, descr_beneficio)
                        if code_res:
                            cclasstrib       = code_res
                            descr_cclasstrib = msg_res
                except Exception:
                    pass

        total_ibs = round(ibs_uf + ibs_mun, 6)
        total_iva = round(total_ibs + cbs, 6)

        base_legal = BASE_LEGAL_MAP.get(regime, "LC 214/2025 — Consulte o regime específico")
        if anexo_beneficio and "Anexo" not in base_legal:
            base_legal = f"{base_legal} ({anexo_beneficio})"

        enriquecidos.append({
            **item,
            "cclasstrib":        cclasstrib,
            "descr_cclasstrib":  descr_cclasstrib,
            "regime_iva":        regime,
            "reducao_pct":       reducao_pct,
            "anexo_beneficio":   anexo_beneficio,
            "descr_beneficio":   descr_beneficio,
            "ibs_uf_pct":        ibs_uf,
            "ibs_mun_pct":       ibs_mun,
            "total_ibs_pct":     total_ibs,
            "cbs_pct":           cbs,
            "total_iva_pct":     total_iva,
            "base_legal":        base_legal,
            "flag_ambiguo":      flag_ambiguo,
            "opcoes_cclasstrib": opcoes_str,
        })

    return enriquecidos


# =============================================================================
# EXPORTAÇÃO EXCEL
# =============================================================================

def gerar_excel_efd(
    df_c180: pd.DataFrame,
    df_c175: pd.DataFrame,
    empresa: Dict,
    modo: str,
    nome_arquivo: str,
    apuracao_pis: Dict,
    apuracao_cofins: Dict,
    receita_bruta: Dict,
) -> bytes:
    """
    Gera planilha Excel com cores PRICETAX (amarelo #FFDD00 / preto #0A0A0A).

    Abas:
    1. NF-e por NCM (C180) — com cClassTrib derivado
    2. NFC-e por CFOP (C175) — agrupado sem NCM individual
    3. Apuração PIS/COFINS
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        st.error("Instale openpyxl: pip install openpyxl")
        return b""

    wb = openpyxl.Workbook()

    AMARELO = "FFFFDD00"
    PRETO   = "FF0A0A0A"
    LARANJA = "FFFF6D00"
    BRANCO  = "FFFFFFFF"
    VERDE   = "FF2E7D32"

    def hdr(ws, row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.fill = PatternFill("solid", fgColor=PRETO)
        c.font = Font(name="Calibri", bold=True, size=10, color=AMARELO)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        return c

    def ttl(ws, row, col, value, merge_to=None):
        c = ws.cell(row=row, column=col, value=value)
        c.fill = PatternFill("solid", fgColor=AMARELO)
        c.font = Font(name="Calibri", bold=True, size=11, color=PRETO)
        c.alignment = Alignment(horizontal="center", vertical="center")
        if merge_to:
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to)
        return c

    empresa_str = (
        f"{empresa.get('nome', '')} | CNPJ: {empresa.get('cnpj', '')} | "
        f"Período: {empresa.get('periodo', '')} a {empresa.get('periodo_fim', '')}"
    )

    # ── Aba 1: NF-e por NCM (C180) ────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "NF-e por NCM (C180)"
    ws1.freeze_panes = "A3"
    ttl(ws1, 1, 1, f"PRICETAX — EFD Contribuições | {empresa_str}", merge_to=12)
    ws1.row_dimensions[1].height = 22

    cols1 = ["NCM", "Descrição", "CFOP", "CST PIS", "Período", "Valor Total (R$)",
             "cClassTrib", "Regime IBS/CBS", "Redução (%)", "Total IVA (%)", "Base Legal", "Ambíguo?", "Opções cClassTrib"]
    for i, c in enumerate(cols1, 1):
        hdr(ws1, 2, i, c)
    ws1.row_dimensions[2].height = 28

    if not df_c180.empty:
        for idx, row in enumerate(df_c180.sort_values("vl_total", ascending=False).itertuples(), 1):
            r = idx + 2
            ambiguo = getattr(row, "flag_ambiguo", False)
            ws1.cell(r, 1,  getattr(row, "ncm", "")).font = Font(name="Calibri", size=9, bold=True)
            ws1.cell(r, 2,  getattr(row, "descricao", "")).font = Font(name="Calibri", size=9)
            ws1.cell(r, 3,  getattr(row, "cfop", "")).font = Font(name="Calibri", size=9)
            ws1.cell(r, 4,  getattr(row, "cst_pis", "")).font = Font(name="Calibri", size=9)
            ws1.cell(r, 5,  f"{getattr(row, 'dt_ini', '')} a {getattr(row, 'dt_fin', '')}").font = Font(name="Calibri", size=9)
            ws1.cell(r, 6,  getattr(row, "vl_total", 0.0)).number_format = 'R$ #,##0.00'
            ws1.cell(r, 6).font = Font(name="Calibri", size=9)
            cc = ws1.cell(r, 7, getattr(row, "cclasstrib", ""))
            cc.font = Font(name="Calibri", size=9, bold=True, color=PRETO if not ambiguo else BRANCO)
            cc.fill = PatternFill("solid", fgColor=LARANJA if ambiguo else AMARELO)
            ws1.cell(r, 8,  getattr(row, "regime_iva", "")).font = Font(name="Calibri", size=9)
            ws1.cell(r, 9,  getattr(row, "reducao_pct", 0)).number_format = '0"%"'
            ws1.cell(r, 9).font = Font(name="Calibri", size=9)
            ws1.cell(r, 10, round(getattr(row, "total_iva_pct", 0.0) * 100, 4)).number_format = '0.0000"%"'
            ws1.cell(r, 10).font = Font(name="Calibri", size=9)
            ws1.cell(r, 11, getattr(row, "base_legal", "")).font = Font(name="Calibri", size=9)
            ws1.cell(r, 12, "SIM" if ambiguo else "NÃO").font = Font(name="Calibri", size=9, bold=ambiguo, color="FF6D00" if ambiguo else VERDE)
            ws1.cell(r, 13, getattr(row, "opcoes_cclasstrib", "")).font = Font(name="Calibri", size=9)

    for col, w in zip("ABCDEFGHIJKLM", [14, 40, 8, 8, 22, 18, 12, 28, 10, 12, 55, 10, 45]):
        ws1.column_dimensions[col].width = w

    # ── Aba 2: NFC-e por CFOP (C175) ─────────────────────────────────────────
    ws2 = wb.create_sheet("NFC-e por CFOP (C175)")
    ws2.freeze_panes = "A3"
    ttl(ws2, 1, 1, f"PRICETAX — NFC-e (C175) | {empresa_str}", merge_to=5)

    cols2 = ["CFOP", "Descrição", "CST PIS", "Valor Total (R$)", "Qtd Notas"]
    for i, c in enumerate(cols2, 1):
        hdr(ws2, 2, i, c)

    if not df_c175.empty:
        for idx, row in enumerate(df_c175.sort_values("vl_total", ascending=False).itertuples(), 1):
            r = idx + 2
            ws2.cell(r, 1, getattr(row, "cfop", "")).font = Font(name="Calibri", size=9, bold=True)
            ws2.cell(r, 2, getattr(row, "descricao", "")).font = Font(name="Calibri", size=9)
            ws2.cell(r, 3, getattr(row, "cst_pis", "")).font = Font(name="Calibri", size=9)
            ws2.cell(r, 4, getattr(row, "vl_total", 0.0)).number_format = 'R$ #,##0.00'
            ws2.cell(r, 4).font = Font(name="Calibri", size=9)
            ws2.cell(r, 5, getattr(row, "qtd_notas", 0)).font = Font(name="Calibri", size=9)

    for col, w in zip("ABCDE", [8, 30, 8, 18, 10]):
        ws2.column_dimensions[col].width = w

    # ── Aba 3: Apuração PIS/COFINS ────────────────────────────────────────────
    ws3 = wb.create_sheet("Apuração PIS-COFINS")
    ttl(ws3, 1, 1, f"PRICETAX — Apuração PIS/COFINS | {empresa_str}", merge_to=3)

    hdr(ws3, 2, 1, "Campo")
    hdr(ws3, 2, 2, "PIS (R$)")
    hdr(ws3, 2, 3, "COFINS (R$)")

    campos_apuracao = [
        ("Contribuição NC do Período",    "vl_tot_cont_nc_per"),
        ("Créditos a Descontar NC",       "vl_cred_desc_nc_per"),
        ("Créditos de Períodos Anteriores", "vl_cred_desc_ant"),
        ("Contribuição Cumulativa",       "vl_tot_cont_cum_per"),
        ("Retenções NC",                  "vl_ret_nc"),
        ("Retenções Cumulativo",          "vl_ret_cum"),
        ("Outras Deduções NC",            "vl_out_ded_nc"),
        ("Contribuição NC a Recolher",    "vl_cont_nc_rec"),
        ("Contribuição Cumulativa a Recolher", "vl_cont_cum_rec"),
        ("TOTAL A RECOLHER",              "vl_tot_cont_rec"),
    ]

    for i, (label, key) in enumerate(campos_apuracao, 3):
        ws3.cell(i, 1, label).font = Font(name="Calibri", size=9, bold=(key == "vl_tot_cont_rec"))
        ws3.cell(i, 2, apuracao_pis.get(key, 0.0)).number_format = 'R$ #,##0.00'
        ws3.cell(i, 2).font = Font(name="Calibri", size=9, bold=(key == "vl_tot_cont_rec"))
        ws3.cell(i, 3, apuracao_cofins.get(key, 0.0)).number_format = 'R$ #,##0.00'
        ws3.cell(i, 3).font = Font(name="Calibri", size=9, bold=(key == "vl_tot_cont_rec"))

    ws3.column_dimensions["A"].width = 40
    ws3.column_dimensions["B"].width = 18
    ws3.column_dimensions["C"].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# =============================================================================
# RENDERIZAÇÃO DA ABA NO STREAMLIT
# =============================================================================

def render_aba_efd_contribuicoes(beneficios_engine=None, consulta_ncm_fn=None):
    """
    Renderiza a aba EFD Contribuições no Streamlit.

    Parâmetros:
        beneficios_engine: instância do BeneficiosFiscaisEngine (do app.py)
        consulta_ncm_fn:   função consulta_ncm do beneficios_fiscais.py
    """
    st.markdown("""
    <div style="background:#1A1A1A;border-left:4px solid #FFDD00;padding:1rem 1.25rem;border-radius:4px;margin-bottom:1.5rem;">
        <div style="font-size:1.1rem;font-weight:700;color:#FFDD00;margin-bottom:0.25rem;">
            EFD Contribuições — cClassTrib IBS/CBS por NCM
        </div>
        <div style="font-size:0.82rem;color:#AAAAAA;">
            Processa arquivos EFD Contribuições (.txt) e deriva automaticamente o cClassTrib IBS/CBS por produto.
            Suporta arquivos com C175 (NFC-e/varejo) e C180 (NF-e modelo 55) — ambos podem coexistir no mesmo arquivo.
        </div>
    </div>
    """, unsafe_allow_html=True)

    arquivo = st.file_uploader(
        "Selecione o arquivo EFD Contribuições (.txt)",
        type=["txt"],
        key="efd_contribuicoes_upload",
        help="Arquivo SPED EFD Contribuições no formato .txt (encoding Latin-1 ou UTF-8)"
    )

    if not arquivo:
        st.info("Aguardando upload do arquivo EFD Contribuições (.txt)...")
        return

    if st.button("Processar EFD Contribuições", type="primary", key="efd_processar_btn"):
        with st.spinner("Lendo e processando o arquivo EFD Contribuições..."):
            try:
                conteudo = arquivo.read().decode("latin-1", errors="replace")
            except Exception as e:
                st.error(f"Erro ao ler o arquivo: {e}")
                return

            resultado = processar_efd_contribuicoes(conteudo)

        # Alertas de processamento
        for alerta in resultado["alertas"]:
            st.warning(f"Aviso: {alerta}")

        empresa = resultado["empresa"]
        modo    = resultado["modo"]

        if modo == "DESCONHECIDO":
            st.error("Nenhum item de saída encontrado. Verifique se o arquivo é um EFD Contribuições válido.")
            return

        # Cabeçalho empresa
        col_emp1, col_emp2, col_emp3 = st.columns(3)
        col_emp1.metric("Empresa", empresa.get("nome", "—")[:30])
        col_emp2.metric("CNPJ", empresa.get("cnpj", "—"))
        col_emp3.metric("Período", f"{empresa.get('periodo', '—')} a {empresa.get('periodo_fim', '—')}")

        # Badge do modo
        modo_labels = {
            "MISTO":    "Modo MISTO: arquivo com NF-e (C180 com NCM) e NFC-e (C175 sem NCM individual).",
            "C180_NFE55": "Modo C180: apenas NF-e modelo 55 com NCM por item.",
            "C175_NFCE":  "Modo C175: apenas NFC-e (cupom fiscal) sem NCM individual por item.",
        }
        st.info(modo_labels.get(modo, f"Modo: {modo}"), icon="ℹ️")

        # ── Seção C180 (NF-e com NCM) ─────────────────────────────────────────
        itens_c180 = resultado["itens_c180"]
        if itens_c180:
            st.markdown("---")
            st.markdown("### NF-e por NCM (C180) — cClassTrib IBS/CBS")
            st.caption(
                "Estes são os itens de NF-e modelo 55 com NCM identificado. "
                "O cClassTrib é derivado automaticamente via base PRICETAX."
            )

            with st.spinner("Derivando cClassTrib por NCM..."):
                itens_enriquecidos = enriquecer_cclasstrib(itens_c180, beneficios_engine, consulta_ncm_fn)

            df_c180 = pd.DataFrame(itens_enriquecidos)

            total_ambiguos  = int(df_c180["flag_ambiguo"].sum()) if "flag_ambiguo" in df_c180.columns else 0
            total_beneficio = int((df_c180["reducao_pct"] > 0).sum()) if "reducao_pct" in df_c180.columns else 0
            total_aliq_zero = int((df_c180["reducao_pct"] == 100).sum()) if "reducao_pct" in df_c180.columns else 0

            k1, k2, k3, k4 = st.columns(4)
            k1.metric("NCMs processados", len(df_c180))
            k2.metric("Com benefício fiscal", total_beneficio)
            k3.metric("Alíquota zero", total_aliq_zero)
            k4.metric("NCMs ambíguos", total_ambiguos,
                      delta="revisão manual" if total_ambiguos > 0 else None,
                      delta_color="inverse")

            if total_ambiguos > 0:
                st.warning(
                    f"{total_ambiguos} NCM(s) com múltiplos cClassTribs possíveis — "
                    "verifique a coluna 'Ambíguo?' e 'Opções cClassTrib' no Excel."
                )

            cols_tela = [c for c in [
                "ncm", "descricao", "cfop", "cst_pis", "vl_total",
                "cclasstrib", "regime_iva", "reducao_pct",
                "total_iva_pct", "base_legal", "flag_ambiguo", "opcoes_cclasstrib"
            ] if c in df_c180.columns]

            st.dataframe(
                df_c180[cols_tela],
                use_container_width=True,
                column_config={
                    "ncm":               st.column_config.TextColumn("NCM"),
                    "descricao":         st.column_config.TextColumn("Descrição"),
                    "cfop":              st.column_config.TextColumn("CFOP"),
                    "cst_pis":           st.column_config.TextColumn("CST PIS"),
                    "vl_total":          st.column_config.NumberColumn("Valor Total (R$)", format="R$ %.2f"),
                    "cclasstrib":        st.column_config.TextColumn("cClassTrib"),
                    "regime_iva":        st.column_config.TextColumn("Regime IBS/CBS"),
                    "reducao_pct":       st.column_config.NumberColumn("Redução (%)", format="%.0f%%"),
                    "total_iva_pct":     st.column_config.NumberColumn("Total IVA (%)", format="%.4f%%"),
                    "base_legal":        st.column_config.TextColumn("Base Legal"),
                    "flag_ambiguo":      st.column_config.CheckboxColumn("Ambíguo?"),
                    "opcoes_cclasstrib": st.column_config.TextColumn("Opções cClassTrib"),
                },
                hide_index=True,
            )
        else:
            df_c180 = pd.DataFrame()

        # ── Seção C175 (NFC-e por CFOP) ───────────────────────────────────────
        itens_c175 = resultado["itens_c175"]
        if itens_c175:
            st.markdown("---")
            st.markdown("### NFC-e por CFOP (C175) — Resumo por CFOP")
            st.caption(
                "NFC-e (cupom fiscal eletrônico) não possui NCM individual por item no EFD Contribuições. "
                "Os valores estão agrupados por CFOP. Para derivação de cClassTrib por NCM das NFC-e, "
                "utilize o Ranking de Saídas SPED (que lê o SPED Fiscal com C170/0200)."
            )

            df_c175 = pd.DataFrame(itens_c175)
            total_nfce = df_c175["vl_total"].sum() if "vl_total" in df_c175.columns else 0.0

            kn1, kn2 = st.columns(2)
            kn1.metric("Total NFC-e", f"R$ {total_nfce:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
            kn2.metric("Notas processadas", resultado.get("total_notas_nfce", 0))

            st.dataframe(
                df_c175[["cfop", "descricao", "cst_pis", "vl_total", "qtd_notas"]],
                use_container_width=True,
                column_config={
                    "cfop":      st.column_config.TextColumn("CFOP"),
                    "descricao": st.column_config.TextColumn("Descrição"),
                    "cst_pis":   st.column_config.TextColumn("CST PIS"),
                    "vl_total":  st.column_config.NumberColumn("Valor Total (R$)", format="R$ %.2f"),
                    "qtd_notas": st.column_config.NumberColumn("Qtd Notas"),
                },
                hide_index=True,
            )
        else:
            df_c175 = pd.DataFrame()

        # ── Apuração PIS/COFINS ────────────────────────────────────────────────
        pis    = resultado.get("apuracao_pis", {})
        cofins = resultado.get("apuracao_cofins", {})

        if pis or cofins:
            st.markdown("---")
            st.markdown("**Apuração PIS/COFINS (M200/M600)**")
            ap1, ap2, ap3, ap4 = st.columns(4)
            _fmt = lambda v: f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            ap1.metric("PIS a recolher",    _fmt(pis.get("vl_tot_cont_rec", 0)))
            ap2.metric("COFINS a recolher", _fmt(cofins.get("vl_tot_cont_rec", 0)))
            ap3.metric("Créditos PIS",      _fmt(pis.get("vl_cred_desc_nc_per", 0)))
            ap4.metric("Créditos COFINS",   _fmt(cofins.get("vl_cred_desc_nc_per", 0)))

        # ── Download Excel ─────────────────────────────────────────────────────
        st.markdown("---")
        excel_bytes = gerar_excel_efd(
            df_c180=df_c180,
            df_c175=df_c175,
            empresa=empresa,
            modo=modo,
            nome_arquivo=arquivo.name,
            apuracao_pis=pis,
            apuracao_cofins=cofins,
            receita_bruta=resultado.get("receita_bruta", {}),
        )
        if excel_bytes:
            st.download_button(
                label="Download Excel — EFD Contribuições cClassTrib IBS/CBS",
                data=excel_bytes,
                file_name=f"efd_contribuicoes_cclasstrib_{arquivo.name.replace('.txt', '')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

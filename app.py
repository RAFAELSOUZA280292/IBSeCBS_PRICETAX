import io
import re
import zipfile
from collections import defaultdict
from pathlib import Path
from typing import Optional, Dict, Any

import pandas as pd
import streamlit as st
import altair as alt

# --------------------------------------------------
# CONFIG GERAL / TEMA PRICETAX
# --------------------------------------------------
st.set_page_config(
    page_title="PRICETAX • IBS/CBS 2026 & Ranking SPED",
    page_icon="💡",
    layout="wide",
)

PRIMARY_YELLOW = "#FFC300"
PRIMARY_BLACK = "#050608"
PRIMARY_CYAN = "#007BFF"  # azul normal

st.markdown(
    f"""
    <style>
    .stApp {{
        background-color: {PRIMARY_BLACK};
        color: #F5F5F5;
        font-family: "Segoe UI", system-ui, -apple-system, BlinkMacSystemFont, sans-serif;
    }}

    /* Layout geral */
    .block-container {{
        padding-top: 1.5rem;
        padding-bottom: 2rem;
        max-width: 1200px;
    }}

    /* Títulos */
    .pricetax-title {{
        font-size: 2.3rem;
        font-weight: 700;
        color: {PRIMARY_YELLOW};
        letter-spacing: 0.04em;
    }}
    .pricetax-subtitle {{
        font-size: 0.95rem;
        color: #CFCFCF;
        margin-top: 0.2rem;
    }}

    /* Cards */
    .pricetax-card {{
        border-radius: 0.9rem;
        padding: 1.1rem 1.3rem;
        background: linear-gradient(135deg, #1C1C1C 0%, #101010 60%, #060608 100%);
        border: 1px solid #333333;
    }}
    .pricetax-card-soft {{
        border-radius: 0.9rem;
        padding: 1.1rem 1.3rem;
        background: #111318;
        border: 1px solid #2B2F3A;
    }}
    .pricetax-card-erro {{
        border-radius: 0.9rem;
        padding: 1.1rem 1.3rem;
        background: #2b1a1a;
        border: 1px solid #ff5656;
    }}

    /* Badges e chips */
    .pricetax-badge {{
        display: inline-block;
        padding: 0.18rem 0.7rem;
        border-radius: 999px;
        background: {PRIMARY_YELLOW};
        color: {PRIMARY_BLACK};
        font-size: 0.72rem;
        font-weight: 700;
        text-transform: uppercase;
        letter-spacing: 0.08em;
    }}
    .pill {{
        display: inline-flex;
        align-items: center;
        gap: 0.3rem;
        padding: 0.18rem 0.7rem;
        border-radius: 999px;
        font-size: 0.78rem;
        font-weight: 500;
        border: 1px solid rgba(255,255,255,0.08);
        background: rgba(15,15,18,0.9);
        color: #EDEDED;
    }}
    .pill-regime {{
        border-color: {PRIMARY_CYAN};
        background: rgba(0,123,255,0.12);
        color: #E5F0FF;
    }}
    .pill-tag {{
        background: rgba(0,0,0,0.4);
    }}

    /* Métricas */
    .pricetax-metric-label {{
        font-size: 0.78rem;
        color: #BBBBBB;
        text-transform: uppercase;
        letter-spacing: 0.05em;
    }}

    /* Tabs */
    .stTabs [data-baseweb="tab-list"] {{
        border-bottom: 1px solid #333333;
    }}
    .stTabs [data-baseweb="tab"] {{
        padding-top: 0.6rem;
        padding-bottom: 0.4rem;
    }}
    .stTabs [aria-selected="true"] p {{
        color: {PRIMARY_YELLOW} !important;
        font-weight: 600;
    }}

    /* Inputs */
    .stTextInput > div > div > input {{
        background-color: #111318;
        color: #FFFFFF;
        border-radius: 0.6rem;
        border: 1px solid #333333;
    }}
    .stFileUploader > label div {{
        color: #DDDDDD;
    }}

    /* Botão primário - azul */
    .stButton>button[kind="primary"] {{
        background-color: {PRIMARY_CYAN};
        color: #ffffff;
        border-radius: 0.6rem;
        border: 1px solid {PRIMARY_CYAN};
        font-weight: 600;
        padding: 0.4rem 1.5rem;
    }}
    .stButton>button[kind="primary"]:hover {{
        background-color: #0069d9;
        border-color: #5fa8ff;
    }}
    </style>
    """,
    unsafe_allow_html=True,
)

# --------------------------------------------------
# UTILITÁRIOS
# --------------------------------------------------
def only_digits(s: Optional[str]) -> str:
    return re.sub(r"\D+", "", s or "")


def to_float_br(s) -> float:
    if s is None:
        return 0.0
    s = str(s).strip()
    if s == "":
        return 0.0
    if s.count(",") == 1 and s.count(".") >= 1:
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0


def competencia_from_dt(dt_ini: str, dt_fin: str) -> str:
    """
    Extrai competência (MM/AAAA) a partir das datas do registro 0000.
    """
    for raw in (dt_ini or "", dt_fin or ""):
        dig = only_digits(raw)
        if len(dig) == 8:
            return f"{dig[2:4]}/{dig[4:8]}"
    return ""


def normalize_cols_upper(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip().upper() for c in df.columns]
    return df


def pct_str(v: float) -> str:
    return f"{v:.2f}".replace(".", ",") + "%"


emoji_sim = "🔵"
emoji_nao = "🔴"


def badge_flag(v):
    v = str(v or "").strip().upper()
    return f"{emoji_sim} SIM" if v == "SIM" else f"{emoji_nao} NÃO"


def regime_label(regime: str) -> str:
    r = (regime or "").upper()
    mapping = {
        "ALIQ_ZERO_CESTA_BASICA_NACIONAL": "Alíquota zero • Cesta Básica Nacional",
        "ALIQ_ZERO_HORTIFRUTI_OVOS": "Alíquota zero • Hortifrúti e Ovos",
        "RED_60_ALIMENTOS": "Redução de 60% • Alimentos",
        "RED_60_ESSENCIALIDADE": "Redução de 60% • Essencialidade",
        "TRIBUTACAO_PADRAO": "Tributação padrão (sem benefício)",
    }
    return mapping.get(r, regime or "Regime não mapeado")


def label_from_sped_header(text: str, default_name: str) -> str:
    """
    Monta rótulo "MM/AAAA - NOME DA EMPRESA" a partir do registro |0000|.
    Se não conseguir, retorna o nome padrão (nome do arquivo).
    """
    try:
        for line in text.splitlines():
            if line.startswith("|0000|"):
                parts = line.split("|")
                dt_ini = parts[4] if len(parts) > 4 else ""
                dt_fin = parts[5] if len(parts) > 5 else ""
                nome = parts[6] if len(parts) > 6 else ""
                comp = competencia_from_dt(dt_ini, dt_fin)
                nome_clean = nome.strip() or default_name
                if comp:
                    return f"{comp} - {nome_clean}"
                return nome_clean
    except Exception:
        pass
    return default_name


# --------------------------------------------------
# CARREGAR BASE TIPI (PLANILHA PRICETAX / MIND7)
# --------------------------------------------------
TIPI_DEFAULT_NAME = "PLANILHA_PRICETAX_REGRAS_REFINADAS.xlsx"
ALT_TIPI_NAME = "TIPI_IBS_CBS_CLASSIFICADA_MIND7.xlsx"


@st.cache_data(show_spinner=False)
def load_tipi_base() -> pd.DataFrame:
    paths = [
        Path(TIPI_DEFAULT_NAME),
        Path.cwd() / TIPI_DEFAULT_NAME,
        Path(ALT_TIPI_NAME),
        Path.cwd() / ALT_TIPI_NAME,
    ]
    try:
        paths.append(Path(__file__).parent / TIPI_DEFAULT_NAME)
        paths.append(Path(__file__).parent / ALT_TIPI_NAME)
    except Exception:
        pass

    df = None
    for p in paths:
        if p.exists():
            df = pd.read_excel(p)
            break

    if df is None:
        return pd.DataFrame()

    df = normalize_cols_upper(df)
    if "NCM" not in df.columns:
        return pd.DataFrame()

    if "NCM_DIG" not in df.columns:
        df["NCM_DIG"] = (
            df["NCM"].astype(str).str.replace(r"\D", "", regex=True).str.zfill(8)
        )

    required = [
        "NCM_DESCRICAO",
        "REGIME_IVA_2026_FINAL",
        "FONTE_LEGAL_FINAL",
        "FLAG_ALIMENTO",
        "FLAG_CESTA_BASICA",
        "FLAG_HORTIFRUTI_OVOS",
        "FLAG_RED_60",
        "FLAG_DEPENDE_DESTINACAO",
        "IBS_UF_TESTE_2026_FINAL",
        "IBS_MUN_TESTE_2026_FINAL",
        "CBS_TESTE_2026_FINAL",
        "CST_IBSCBS",
        "ALERTA_APP",
        "OBS_ALIMENTO",
        "OBS_DESTINACAO",
        "OBS_REGIME_ESPECIAL",
        "FLAG_IMPOSTO_SELETIVO",
    ]
    for c in required:
        if c not in df.columns:
            df[c] = ""

    return df


def buscar_ncm(df: pd.DataFrame, ncm_raw: str):
    n = only_digits(ncm_raw)
    if len(n) != 8 or df.empty:
        return None
    row = df.loc[df["NCM_DIG"] == n]
    return None if row.empty or row.isnull().all().all() else row.iloc[0]


df_tipi = load_tipi_base()

# --------------------------------------------------
# CARREGAR BASE CLASSIFICAÇÃO TRIBUTÁRIA (por código)
# --------------------------------------------------
CLASSIF_NAME = "classificacao_tributaria.xlsx"


@st.cache_data(show_spinner=False)
def load_classificacao_base() -> pd.DataFrame:
    paths = [
        Path(CLASSIF_NAME),
        Path.cwd() / CLASSIF_NAME,
    ]
    try:
        paths.append(Path(__file__).parent / CLASSIF_NAME)
    except Exception:
        pass

    df = None
    for p in paths:
        if p.exists():
            df = pd.read_excel(p, sheet_name="Classificação Tributária")
            break

    if df is None:
        return pd.DataFrame()

    df = df.copy()
    return df


df_class = load_classificacao_base()


@st.cache_data(show_spinner=False)
def build_cclasstrib_code_index(df_class_: pd.DataFrame) -> Dict[str, Dict[str, str]]:
    index: Dict[str, Dict[str, str]] = {}
    if df_class_.empty:
        return index

    col_cod = "Código da Classificação Tributária"
    for code, grp in df_class_.groupby(col_cod):
        if pd.isna(code):
            continue
        g = grp.copy()

        # Preferir NFe = "Sim"
        if "NFe" in g.columns:
            g_pref = g[g["NFe"].astype(str).str.lower() == "sim"]
            if not g_pref.empty:
                g = g_pref

        # Garantir colunas
        for col in [
            "Tributação Regular",
            "Redução de Alíquota",
            "Transferência de Crédito",
            "Diferimento",
        ]:
            if col not in g.columns:
                g[col] = ""

        # Preferir cenário regular sem redução/sem diferimento/sem transf de crédito
        mask_reg = (
            (g["Tributação Regular"].astype(str).str.lower() == "sim")
            & (g["Redução de Alíquota"].astype(str).str.lower() == "não")
            & (g["Transferência de Crédito"].astype(str).str.lower() == "não")
            & (g["Diferimento"].astype(str).str.lower() == "não")
        )
        g_reg = g[mask_reg]
        if not g_reg.empty:
            g = g_reg

        row = g.iloc[0]
        code_str = str(code).strip()
        index[code_str] = {
            "COD_CLASS": code_str,
            "DESC_CLASS": str(
                row.get("Descrição do Código da Classificação Tributária", "")
            ).strip(),
            "TIPO_ALIQUOTA": str(row.get("Tipo de Alíquota", "")).strip(),
            "TRIB_REG": str(row.get("Tributação Regular", "")).strip(),
            "RED_ALIQ": str(row.get("Redução de Alíquota", "")).strip(),
            "TRANSF_CRED": str(row.get("Transferência de Crédito", "")).strip(),
            "DIFERIMENTO": str(row.get("Diferimento", "")).strip(),
            "MONOFASICA": str(row.get("Monofásica", "")).strip(),
            "URL_LEI": str(row.get("Url da Legislação", "")).strip(),
        }

    return index


cclasstrib_code_index = build_cclasstrib_code_index(df_class)


def get_class_info_by_code(cclass_code: str) -> Optional[Dict[str, str]]:
    if not cclass_code:
        return None
    return cclasstrib_code_index.get(str(cclass_code).strip())


# --------------------------------------------------
# REGRAS PADRÃO CFOP → cClassTrib (OPERAÇÕES "NORMAIS")
# --------------------------------------------------
CFOP_CCLASSTRIB_MAP = {
    # Vendas internas
    "5101": "000001",
    "5102": "000001",
    "5103": "000001",
    "5104": "000001",
    "5109": "000001",
    "5110": "000001",
    "5116": "000001",
    "5201": "000001",
    "5403": "000001",
    "5405": "000001",
    "5411": "000001",

    # Vendas interestaduais
    "6101": "000001",
    "6102": "000001",
    "6103": "000001",
    "6104": "000001",
    "6107": "000001",
    "6108": "000001",
    "6109": "000001",
    "6110": "000001",
    "6116": "000001",
    "6201": "000001",
    "6403": "000001",
    "6405": "000001",
    "6411": "000001",

    # Exportação
    "7101": "000001",
    "7102": "000001",

    # Saídas não onerosas genéricas – sem tributação
    "5901": "410999",
    "6901": "410999",
    "5902": "410999",
    "6902": "410999",
    "5910": "410999",
    "6910": "410999",
    "5927": "410999",
    "6927": "410999",
    "5916": "410999",
    "6916": "410999",
    "5959": "410999",
    "6949": "410999",

    # Saída onerosa – faturamento p/ entrega futura
    "5922": "000001",
    "6922": "000001",
}


def guess_cclasstrib(cst: Any, cfop: Any, regime_iva: str) -> tuple[str, str]:
    """
    Sugere um cClassTrib padrão a partir do CFOP e CST
    para operações de venda não especiais.

    Retorna (codigo_cClassTrib, mensagem_explicativa).
    """
    cst_clean = re.sub(r"\D+", "", str(cst or ""))
    cfop_clean = re.sub(r"\D+", "", str(cfop or ""))

    if not cfop_clean:
        return "", "Informe o CFOP da operação de venda para sugerir o cClassTrib padrão."

    # 1) Regra fixa via mapa
    if cfop_clean in CFOP_CCLASSTRIB_MAP:
        code = CFOP_CCLASSTRIB_MAP[cfop_clean]
        msg = (
            f"Regra padrão PRICETAX: CFOP {cfop_clean} → "
            f"cClassTrib {code} (conforme matriz PRICETAX)."
        )
        return code, msg

    # 2) Saída (5, 6 ou 7) com CST de tributação "normal" → 000001
    if cfop_clean[0] in ("5", "6", "7") and cst_clean in {"000", "200", "201", "202", "900"}:
        code = "000001"
        msg = (
            f"Regra genérica: CFOP {cfop_clean} é saída tributada padrão "
            f"→ cClassTrib {code} (tributação regular). "
            "Revise apenas se for operação especial (doação, brinde, bonificação, remessa técnica etc.)."
        )
        return code, msg

    # 3) Não consegui sugerir nada com segurança
    return "", (
        "Não foi possível localizar um cClassTrib padrão para o CFOP informado. "
        "Provável operação especial (devolução, bonificação, remessa, teste, garantia etc.) – revisar manualmente."
    )


# --------------------------------------------------
# PROCESSADOR SPED – RANKING DE SAÍDAS (C100/C170, CFOP 5/6/7)
# --------------------------------------------------
def process_sped_file(file_content: str) -> pd.DataFrame:
    """
    Processa o conteúdo do arquivo SPED PIS/COFINS para extrair dados de vendas.
    Retorna um DataFrame com colunas: NCM, DESCRICAO, CFOP, VALOR_TOTAL_VENDAS.
    """
    produtos = {}
    documentos = {}
    itens_venda = []

    cfop_saida_pattern = re.compile(r"^[567]\d{3}$")
    current_doc_key = None

    try:
        file_stream = io.StringIO(file_content)

        for line in file_stream:
            fields = line.strip().split("|")
            if not fields or len(fields) < 2:
                continue

            registro = fields[1]

            if registro == "0200":
                # |0200|COD_ITEM|DESCR_ITEM|...|COD_NCM(8)|...
                if len(fields) >= 9:
                    cod_item = fields[2]
                    descr_item = fields[3]
                    cod_ncm = fields[8]
                    produtos[cod_item] = {"NCM": cod_ncm, "DESCR_ITEM": descr_item}

            elif registro == "C100":
                ind_oper = fields[2] if len(fields) > 2 else ""
                if ind_oper == "1":  # Saída
                    chv_nfe = fields[9] if len(fields) > 9 else ""
                    ser = fields[6] if len(fields) > 6 else ""
                    num_doc = fields[7] if len(fields) > 7 else ""

                    if chv_nfe:
                        current_doc_key = chv_nfe
                    elif ser and num_doc:
                        current_doc_key = f"{ser}-{num_doc}"
                    else:
                        current_doc_key = None

                    if current_doc_key:
                        documentos[current_doc_key] = {"IND_OPER": ind_oper}
                else:
                    current_doc_key = None

            elif (
                registro == "C170"
                and current_doc_key
                and documentos.get(current_doc_key, {}).get("IND_OPER") == "1"
            ):
                # |C170|NUM_ITEM|COD_ITEM(3)|DESCR_COMPL|QTD|UNID|VL_ITEM(7)|...|CFOP(11)|...
                if len(fields) >= 12:
                    cod_item = fields[3]
                    vl_item_str = fields[7].replace(",", ".")
                    cfop = fields[11]

                    try:
                        vl_item = float(vl_item_str)
                    except ValueError:
                        continue

                    if cfop_saida_pattern.match(cfop):
                        itens_venda.append(
                            {
                                "COD_ITEM": cod_item,
                                "VL_ITEM": vl_item,
                                "CFOP": cfop,
                                "DOC_KEY": current_doc_key,
                            }
                        )

            elif registro in ("C190", "C300", "D100", "E100"):
                current_doc_key = None

    except Exception as e:
        st.error(f"Erro ao processar o arquivo: {e}")
        return pd.DataFrame()

    ranking_vendas = defaultdict(
        lambda: {"NCM": "", "DESCR_ITEM": "", "CFOP": "", "TOTAL_VENDAS": 0.0}
    )

    for item in itens_venda:
        cod_item = item["COD_ITEM"]
        vl_item = item["VL_ITEM"]
        cfop = item["CFOP"]

        produto_info = produtos.get(cod_item)
        if produto_info:
            ncm = produto_info["NCM"]
            descr_item = produto_info["DESCR_ITEM"]

            chave = (ncm, descr_item, cfop)
            ranking_vendas[chave]["NCM"] = ncm
            ranking_vendas[chave]["DESCR_ITEM"] = descr_item
            ranking_vendas[chave]["CFOP"] = cfop
            ranking_vendas[chave]["TOTAL_VENDAS"] += vl_item

    relatorio = []
    for chave, dados in ranking_vendas.items():
        relatorio.append(
            {
                "NCM": dados["NCM"],
                "DESCRICAO": dados["DESCR_ITEM"],
                "VALOR_TOTAL_VENDAS": dados["TOTAL_VENDAS"],
                "CFOP": dados["CFOP"],
            }
        )

    if not relatorio:
        return pd.DataFrame(
            columns=["NCM", "DESCRICAO", "CFOP", "VALOR_TOTAL_VENDAS"]
        )

    df = pd.DataFrame(relatorio)
    df = df.sort_values("VALOR_TOTAL_VENDAS", ascending=False).reset_index(drop=True)
    return df


# --------------------------------------------------
# CABEÇALHO / TABS
# --------------------------------------------------
st.markdown(
    """
    <div class="pricetax-title">PRICETAX • IBS/CBS 2026 & Ranking SPED</div>
    <div class="pricetax-subtitle">
        Ferramenta de apoio para parametrização do XML NFe com base no NCM, CFOP e nas regras IBS/CBS 2026, além de ranking de saídas a partir do SPED PIS/COFINS.
    </div>
    """,
    unsafe_allow_html=True,
)

tabs = st.tabs(
    [
        "🔍 Consulta NCM → IBS/CBS 2026 + cClassTrib",
        "📊 Ranking de Saídas (SPED PIS/COFINS → IBS/CBS + cClassTrib)",
    ]
)

# --------------------------------------------------
# ABA 1 – CONSULTA NCM (com cClassTrib via CFOP)
# --------------------------------------------------
with tabs[0]:
    st.markdown(
        """
        <div class="pricetax-card">
            <span class="pricetax-badge">Consulta por NCM</span>
            <div style="margin-top:0.5rem;font-size:0.9rem;color:#DDDDDD;">
                Use este painel como referência para parametrizar o item no ERP e no XML:
                <br><br>
                • Informe o <b>NCM</b> do produto e o <b>CFOP de venda</b> atualmente utilizado;<br>
                • A ferramenta retorna o regime de IVA, as alíquotas IBS/CBS simuladas para 2026;<br>
                • Sugere o <b>cClassTrib</b> padrão para NFe, a partir do CFOP informado;<br>
                • Expõe os principais campos para configuração do XML (pIBS, pCBS, cClassTrib).
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    col1, col2, col3 = st.columns([3, 1.4, 1])
    with col1:
        ncm_input = st.text_input(
            "NCM do produto",
            placeholder="Ex.: 16023220 ou 16.02.32.20",
            help="Informe o NCM completo (8 dígitos), com ou sem pontos.",
        )
    with col2:
        cfop_input = st.text_input(
            "CFOP da venda (atual)",
            placeholder="Ex.: 5102",
            max_chars=4,
            help="CFOP utilizado hoje na venda do produto (quatro dígitos).",
        )
    with col3:
        st.write("")
        consultar = st.button("Consultar parâmetros", type="primary")

    if consultar and ncm_input.strip():
        row = buscar_ncm(df_tipi, ncm_input)

        if row is None:
            st.markdown(
                f"""
                <div class="pricetax-card-erro" style="margin-top:0.8rem;">
                    NCM informado: <b>{ncm_input}</b><br>
                    Não localizamos esse NCM na base PRICETAX. Revise o código ou a planilha de referência.
                </div>
                """,
                unsafe_allow_html=True,
            )
        else:
            ncm_fmt = row["NCM_DIG"]
            desc = row["NCM_DESCRICAO"]
            regime = row["REGIME_IVA_2026_FINAL"]
            fonte = row["FONTE_LEGAL_FINAL"]
            flag_cesta = row["FLAG_CESTA_BASICA"]
            flag_hf = row["FLAG_HORTIFRUTI_OVOS"]
            flag_red = row["FLAG_RED_60"]
            flag_alim = row["FLAG_ALIMENTO"]
            flag_dep = row["FLAG_DEPENDE_DESTINACAO"]
            ibs_uf = to_float_br(row["IBS_UF_TESTE_2026_FINAL"])
            ibs_mun = to_float_br(row["IBS_MUN_TESTE_2026_FINAL"])
            cbs = to_float_br(row["CBS_TESTE_2026_FINAL"])
            total_iva = ibs_uf + ibs_mun + cbs
            cst_ibscbs = row.get("CST_IBSCBS", "")

            # Sugere cClassTrib a partir do CFOP informado
            cclastrib_code, cclastrib_msg = guess_cclasstrib(
                cst=cst_ibscbs, cfop=cfop_input, regime_iva=str(regime or "")
            )
            class_info = get_class_info_by_code(cclastrib_code)

            # Header do produto
            st.markdown(
                f"""
                <div class="pricetax-card" style="margin-top:1rem;">
                    <div style="font-size:1.05rem;font-weight:600;color:{PRIMARY_YELLOW};">
                        NCM {ncm_fmt} – {desc}
                    </div>
                    <div style="margin-top:0.5rem;display:flex;flex-wrap:wrap;gap:0.4rem;">
                        <span class="pill pill-regime">{regime_label(regime)}</span>
                        <span class="pill pill-tag">Cesta Básica: {badge_flag(flag_cesta)}</span>
                        <span class="pill pill-tag">Hortifrúti/Ovos: {badge_flag(flag_hf)}</span>
                        <span class="pill pill-tag">Redução 60%: {badge_flag(flag_red)}</span>
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )

            # Métricas de alíquotas
            st.markdown(
                f"""
                <div class="pricetax-card" style="margin-top:1rem;display:flex;gap:2rem;flex-wrap:wrap;">
                    <div>
                        <div class="pricetax-metric-label">IBS 2026 (UF + Município)</div>
                        <div style="font-size:2.2rem;color:{PRIMARY_YELLOW};">{pct_str(ibs_uf + ibs_mun)}</div>
                    </div>
                    <div>
                        <div class="pricetax-metric-label">CBS 2026</div>
                        <div style="font-size:2.2rem;color:{PRIMARY_YELLOW};">{pct_str(cbs)}</div>
                    </div>
                    <div>
                        <div class="pricetax-metric-label">Carga Total IVA 2026</div>
                        <div style="font-size:2.2rem;color:{PRIMARY_YELLOW};">{pct_str(total_iva)}</div>
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )

            st.subheader("Parâmetros de classificação", divider="gray")
            c1, c2, c3, c4 = st.columns(4)
            with c1:
                st.markdown("**Produto é alimento?**")
                st.markdown(
                    f"<span style='color:{PRIMARY_YELLOW};font-weight:600;'>{badge_flag(flag_alim)}</span>",
                    unsafe_allow_html=True,
                )
            with c2:
                st.markdown("**Cesta Básica Nacional?**")
                st.markdown(
                    f"<span style='color:{PRIMARY_YELLOW};font-weight:600;'>{badge_flag(flag_cesta)}</span>",
                    unsafe_allow_html=True,
                )
            with c3:
                st.markdown("**Hortifrúti / Ovos?**")
                st.markdown(
                    f"<span style='color:{PRIMARY_YELLOW};font-weight:600;'>{badge_flag(flag_hf)}</span>",
                    unsafe_allow_html=True,
                )
            with c4:
                st.markdown("**Depende de destinação?**")
                st.markdown(
                    f"<span style='color:{PRIMARY_YELLOW};font-weight:600;'>{badge_flag(flag_dep)}</span>",
                    unsafe_allow_html=True,
                )

            # Bloco XML / cClassTrib
            st.subheader("Parâmetros para XML 2026 – NFe (venda padrão)", divider="gray")
            col_xml1, col_xml2, col_xml3 = st.columns(3)

            with col_xml1:
                st.markdown("**CST IBS/CBS (venda)**")
                st.markdown(
                    f"<span style='color:{PRIMARY_YELLOW};font-weight:600;'>{cst_ibscbs or '—'}</span>",
                    unsafe_allow_html=True,
                )
                st.markdown("**CFOP informado (venda)**")
                st.markdown(
                    f"<span style='color:{PRIMARY_YELLOW};font-weight:600;'>{(cfop_input or '—').strip()}</span>",
                    unsafe_allow_html=True,
                )

            with col_xml2:
                st.markdown("**cClassTrib sugerido (venda)**")
                if cclastrib_code:
                    desc_class = class_info["DESC_CLASS"] if class_info else ""
                    st.markdown(
                        f"<span style='color:{PRIMARY_YELLOW};font-weight:700;'>"
                        f"{cclastrib_code} – {desc_class}"
                        f"</span>",
                        unsafe_allow_html=True,
                    )
                else:
                    st.markdown(
                        f"<span style='color:{PRIMARY_YELLOW};font-weight:700;'>—</span>",
                        unsafe_allow_html=True,
                    )

                st.markdown("**Tipo de alíquota (cClassTrib)**")
                tipo_aliq = class_info["TIPO_ALIQUOTA"] if class_info else "—"
                st.markdown(tipo_aliq)

            with col_xml3:
                st.markdown("**Imposto Seletivo (IS)**")
                flag_is = row.get("FLAG_IMPOSTO_SELETIVO", "")
                st.markdown(
                    f"<span style='color:{PRIMARY_YELLOW};font-weight:600;'>{badge_flag(flag_is)}</span>",
                    unsafe_allow_html=True,
                )
                if class_info:
                    st.markdown("**Cenário da classificação**")
                    st.markdown(
                        "- Tributação Regular: **{}**  \n"
                        "- Redução de Alíquota: **{}**  \n"
                        "- Transferência de Crédito: **{}**  \n"
                        "- Diferimento: **{}**  \n"
                        "- Monofásica: **{}**".format(
                            class_info.get("TRIB_REG") or "—",
                            class_info.get("RED_ALIQ") or "—",
                            class_info.get("TRANSF_CRED") or "—",
                            class_info.get("DIFERIMENTO") or "—",
                            class_info.get("MONOFASICA") or "—",
                        )
                    )

            st.markdown("**Alíquotas para parametrização no XML (pIBS / pCBS)**")
            st.markdown(
                f"- pIBS (UF): **{pct_str(ibs_uf)}**  \n"
                f"- pIBS (Município): **{pct_str(ibs_mun)}**  \n"
                f"- pCBS: **{pct_str(cbs)}**  \n"
                f"- pIVA Total: **{pct_str(total_iva)}**"
            )

            # Mensagem explicando a regra aplicada
            if cfop_input:
                st.markdown(
                    f"**Regra sugerida para cClassTrib (CFOP {cfop_input.strip()}):** {cclastrib_msg}"
                )
            else:
                st.markdown(
                    "**Regra cClassTrib:** informe o CFOP da operação de venda para sugerirmos o cClassTrib padrão."
                )

            st.markdown("---")

            def clean_txt(v):
                s = str(v or "").strip()
                return "" if s.lower() == "nan" else s

            alerta_fmt = clean_txt(row.get("ALERTA_APP"))
            obs_alim = clean_txt(row.get("OBS_ALIMENTO"))
            obs_dest = clean_txt(row.get("OBS_DESTINACAO"))
            reg_extra = clean_txt(row.get("OBS_REGIME_ESPECIAL"))

            if "RED_60" in (regime or "").upper():
                if not alerta_fmt:
                    alerta_fmt = (
                        "Redução de 60% aplicada; conferir aderência ao segmento e às condições legais."
                    )
                if not reg_extra:
                    reg_extra = (
                        "Ano teste 2026 – IBS 0,1% (UF) e CBS 0,9%. "
                        "Carga reduzida em 60% conforme regras de essencialidade/alimentos."
                    )

            st.markdown(f"**Base legal considerada (TIPI/PRICETAX):** {fonte or '—'}")
            st.markdown(f"**Alerta PRICETAX:** {alerta_fmt or '—'}")
            st.markmarkdown(f"**Observação sobre alimentos:** {obs_alim or '—'}")
            st.markdown(f"**Observação sobre destinação:** {obs_dest or '—'}")
            st.markdown(
                f"**Regime especial / observações adicionais:** {reg_extra or '—'}"
            )

# --------------------------------------------------
# ABA 2 – RANKING DE SAÍDAS (SPED → IBS/CBS + cClassTrib)
# --------------------------------------------------
with tabs[1]:
    st.markdown(
        """
        <div class="pricetax-card">
            <span class="pricetax-badge">Ranking de vendas – SPED PIS/COFINS</span>
            <div style="margin-top:0.5rem;font-size:0.9rem;color:#DDDDDD;">
                Utilize este painel para identificar os itens mais relevantes na receita e preparar a base
                para IBS/CBS 2026:
                <br><br>
                • Importa arquivos SPED PIS/COFINS (<b>.txt</b> ou <b>.zip</b>);<br>
                • Lê o Bloco C (C100/C170) e considera apenas saídas (IND_OPER = 1);<br>
                • Consolida vendas por NCM, descrição do item e CFOP (5.xxx, 6.xxx, 7.xxx);<br>
                • Cruza automaticamente com a TIPI IBS/CBS PRICETAX 2026;<br>
                • Sugere o <b>cClassTrib</b> para cada combinação NCM + CFOP;<br>
                • Gera um ranking exportável em Excel, pronto para trabalho em ERP e BI.
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    uploaded_rank = st.file_uploader(
        "Arquivos SPED PIS/COFINS (.txt ou .zip)",
        type=["txt", "zip"],
        accept_multiple_files=True,
        key="sped_upload_rank",
        help="Você pode selecionar um ou vários arquivos. Arquivos .zip são descompactados automaticamente.",
    )

    if uploaded_rank:
        if st.button("Processar SPED e gerar ranking", type="primary"):
            df_list = []
            total_files = len(uploaded_rank)
            progress_bar = st.progress(0)
            status_text = st.empty()

            with st.spinner("Processando arquivos SPED..."):
                for idx, up in enumerate(uploaded_rank, start=1):
                    nome = up.name

                    if nome.lower().endswith(".zip"):
                        z_bytes = up.read()
                        with zipfile.ZipFile(io.BytesIO(z_bytes), "r") as z:
                            for info in z.infolist():
                                if info.filename.lower().endswith(".txt"):
                                    status_text.markdown(
                                        f"**Processando arquivo {idx}/{total_files}:** `{info.filename}`"
                                    )
                                    conteudo = z.open(info).read()
                                    try:
                                        texto = conteudo.decode("latin-1")
                                    except UnicodeDecodeError:
                                        texto = conteudo.decode(
                                            "utf-8", errors="ignore"
                                        )

                                    df_rank = process_sped_file(texto)
                                    if not df_rank.empty:
                                        label = label_from_sped_header(
                                            texto, info.filename
                                        )
                                        df_rank.insert(0, "ARQUIVO", label)
                                        df_list.append(df_rank)
                    else:
                        status_text.markdown(
                            f"**Processando arquivo {idx}/{total_files}:** `{nome}`"
                        )
                        conteudo = up.read()
                        try:
                            texto = conteudo.decode("latin-1")
                        except UnicodeDecodeError:
                            texto = conteudo.decode("utf-8", errors="ignore")

                        df_rank = process_sped_file(texto)
                        if not df_rank.empty:
                            label = label_from_sped_header(texto, nome)
                            df_rank.insert(0, "ARQUIVO", label)
                            df_list.append(df_rank)

                    progress_bar.progress(idx / total_files)

            status_text.empty()
            progress_bar.empty()

            if not df_list:
                st.error(
                    "Nenhuma nota fiscal de saída com CFOP 5.xxx, 6.xxx ou 7.xxx foi encontrada nos arquivos enviados."
                )
            else:
                df_total = pd.concat(df_list, ignore_index=True)

                # CRUZAMENTO COM TIPI IBS/CBS
                if df_tipi.empty:
                    st.warning(
                        "Base TIPI IBS/CBS 2026 não carregada. O ranking será exibido sem os campos de IBS/CBS/cClassTrib."
                    )
                else:
                    df_total["NCM_DIG"] = (
                        df_total["NCM"]
                        .astype(str)
                        .str.replace(r"\D", "", regex=True)
                        .str.zfill(8)
                    )

                    cols_tipi_merge = [
                        "NCM_DIG",
                        "REGIME_IVA_2026_FINAL",
                        "IBS_UF_TESTE_2026_FINAL",
                        "IBS_MUN_TESTE_2026_FINAL",
                        "CBS_TESTE_2026_FINAL",
                        "CST_IBSCBS",
                    ]
                    for c in cols_tipi_merge:
                        if c not in df_tipi.columns:
                            df_tipi[c] = ""

                    df_total = df_total.merge(
                        df_tipi[cols_tipi_merge],
                        on="NCM_DIG",
                        how="left",
                    )

                    df_total["IBS_UF_2026"] = df_total[
                        "IBS_UF_TESTE_2026_FINAL"
                    ].apply(to_float_br)
                    df_total["IBS_MUN_2026"] = df_total[
                        "IBS_MUN_TESTE_2026_FINAL"
                    ].apply(to_float_br)
                    df_total["CBS_2026"] = df_total["CBS_TESTE_2026_FINAL"].apply(
                        to_float_br
                    )
                    df_total["ALIQ_IVA_TOTAL_2026"] = (
                        df_total["IBS_UF_2026"]
                        + df_total["IBS_MUN_2026"]
                        + df_total["CBS_2026"]
                    )

                    # Sugere cClassTrib por CFOP/CST
                    def apply_guess(row):
                        code, msg = guess_cclasstrib(
                            cst=row.get("CST_IBSCBS", ""),
                            cfop=row.get("CFOP", ""),
                            regime_iva=row.get("REGIME_IVA_2026_FINAL", ""),
                        )
                        return pd.Series([code, msg])

                    df_total[["CCLASSTRIB_SUGERIDO", "CCLASSTRIB_MSG"]] = df_total.apply(
                        apply_guess, axis=1
                    )

                    # Descrição / tipo da classificação, se existir código
                    def map_desc(code):
                        info = get_class_info_by_code(code)
                        return info["DESC_CLASS"] if info else ""

                    def map_tipo(code):
                        info = get_class_info_by_code(code)
                        return info["TIPO_ALIQUOTA"] if info else ""

                    df_total["DESC_CCLASSTRIB"] = df_total["CCLASSTRIB_SUGERIDO"].apply(
                        map_desc
                    )
                    df_total["TIPO_ALIQUOTA_CCLASSTRIB"] = df_total[
                        "CCLASSTRIB_SUGERIDO"
                    ].apply(map_tipo)

                df_total = df_total.sort_values(
                    "VALOR_TOTAL_VENDAS", ascending=False
                ).reset_index(drop=True)

                # VISUALIZAÇÃO – preparando DF de tela
                df_vis = df_total.copy()
                df_vis["VALOR_TOTAL_VENDAS"] = df_vis["VALOR_TOTAL_VENDAS"].apply(
                    lambda v: f"{v:,.2f}"
                    .replace(",", "X")
                    .replace(".", ",")
                    .replace("X", ".")
                )

                for col in [
                    "IBS_UF_2026",
                    "IBS_MUN_2026",
                    "CBS_2026",
                    "ALIQ_IVA_TOTAL_2026",
                ]:
                    if col in df_vis.columns:
                        df_vis[col] = df_vis[col].apply(
                            lambda v: pct_str(v) if pd.notnull(v) else ""
                        )

                # Organiza colunas principais para o usuário
                preferred_cols = [
                    "ARQUIVO",
                    "NCM",
                    "DESCRICAO",
                    "CFOP",
                    "VALOR_TOTAL_VENDAS",
                    "IBS_UF_2026",
                    "IBS_MUN_2026",
                    "CBS_2026",
                    "ALIQ_IVA_TOTAL_2026",
                    "CST_IBSCBS",
                    "CCLASSTRIB_SUGERIDO",
                    "DESC_CCLASSTRIB",
                    "TIPO_ALIQUOTA_CCLASSTRIB",
                ]
                other_cols = [c for c in df_vis.columns if c not in preferred_cols]
                df_vis = df_vis[preferred_cols + other_cols]

                st.success("Processamento concluído.")
                st.markdown("---")

                # Exportação Excel formatada
                def to_excel(df: pd.DataFrame) -> bytes:
                    buf = io.BytesIO()
                    with pd.ExcelWriter(buf, engine="openpyxl") as w:
                        sheet_name = "RANKING_SAIDAS_2026"
                        df.to_excel(
                            w, index=False, sheet_name=sheet_name
                        )
                        ws = w.sheets[sheet_name]

                        # Formatação básica do cabeçalho
                        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

                        header_font = Font(bold=True, color="FFFFFF")
                        header_fill = PatternFill("solid", fgColor="1F2933")
                        thin = Side(border_style="thin", color="444444")
                        border = Border(top=thin, left=thin, right=thin, bottom=thin)

                        for cell in ws[1]:
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.border = border
                            cell.alignment = Alignment(vertical="center")

                        # Congela cabeçalho
                        ws.freeze_panes = "A2"

                        # Ajusta largura das colunas
                        for col_cells in ws.columns:
                            max_length = 0
                            col_letter = col_cells[0].column_letter
                            for cell in col_cells:
                                try:
                                    value = str(cell.value) if cell.value is not None else ""
                                    if len(value) > max_length:
                                        max_length = len(value)
                                except Exception:
                                    pass
                            adjusted_width = min(max_length + 2, 50)
                            ws.column_dimensions[col_letter].width = adjusted_width

                        # Formato monetário para valor total de vendas
                        for cell in ws[1]:
                            if cell.value == "VALOR_TOTAL_VENDAS":
                                col_letter = cell.column_letter
                                for data_cell in ws[col_letter][1:]:
                                    data_cell.number_format = "R$ #,##0.00"

                    buf.seek(0)
                    return buf.read()

                st.download_button(
                    "📥 Baixar Ranking de Saídas 2026 (Excel formatado)",
                    data=to_excel(df_total),
                    file_name="PRICETAX_Ranking_Saidas_Sped_2026.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

                st.markdown(
                    "### Ranking de saídas – visão IBS/CBS 2026 + cClassTrib"
                )
                st.dataframe(df_vis, use_container_width=True)

                total_vendas = df_total["VALOR_TOTAL_VENDAS"].sum()
                total_vendas_fmt = (
                    f"{total_vendas:,.2f}"
                    .replace(",", "X")
                    .replace(".", ",")
                    .replace("X", ".")
                )

                st.markdown(
                    f"""
                    <div class="pricetax-card-soft" style="margin-top:1rem;">
                        <div style="font-size:1rem;color:{PRIMARY_YELLOW};font-weight:600;">Resumo da análise</div>
                        <div style="margin-top:0.4rem;font-size:0.9rem;color:#E0E0E0;">
                            • Total geral de vendas (CFOP 5/6/7): <b>R$ {total_vendas_fmt}</b><br>
                            • Arquivos SPED analisados: <b>{total_files}</b><br>
                            • Ranking consolidado por NCM + descrição + CFOP, já com visão IBS/CBS 2026 e cClassTrib sugerido.<br>
                        </div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

                st.markdown(
                    "### TOP 10 – Distribuição percentual por NCM (vendas de saída)"
                )
                df_top10 = (
                    df_total.groupby("NCM")["VALOR_TOTAL_VENDAS"]
                    .sum()
                    .sort_values(ascending=False)
                    .head(10)
                )

                if not df_top10.empty:
                    df_top10 = df_top10.reset_index()
                    df_top10.rename(
                        columns={
                            "NCM": "NCM",
                            "VALOR_TOTAL_VENDAS": "VALOR_TOTAL_VENDAS",
                        },
                        inplace=True,
                    )

                    chart = (
                        alt.Chart(df_top10)
                        .mark_arc(innerRadius=60)
                        .encode(
                            theta="VALOR_TOTAL_VENDAS:Q",
                            color="NCM:N",
                            tooltip=["NCM:N", "VALOR_TOTAL_VENDAS:Q"],
                        )
                        .properties(
                            width=500,
                            height=400,
                            title="TOP 10 – Percentual por NCM (Vendas de Saída)",
                        )
                    )

                    st.altair_chart(chart, use_container_width=True)
                else:
                    st.info(
                        "Não há dados suficientes para montar o gráfico TOP 10 por NCM."
                    )
    else:
        st.info(
            "Selecione um ou mais arquivos SPED PIS/COFINS para gerar o ranking de saídas."
        )
